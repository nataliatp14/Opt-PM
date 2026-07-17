import streamlit as st
import pandas as pd
import numpy as np
import pydeck as pdk
import altair as alt
import requests
import hashlib
import math
import re
import os
from io import BytesIO
from pathlib import Path
from ortools.constraint_solver import routing_enums_pb2, pywrapcp

# ============================================================
# CONFIGURACIÓN GENERAL
# ============================================================
st.set_page_config(
    page_title="Baseline + Optimizador PM",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
.block-container{    padding-top:2.5rem;    padding-bottom:1rem;}
[data-testid="stHeader"]{    height:3rem;}
.main-title{font-size:2.4rem; font-weight:850; margin-top:0.8rem; margin-bottom:0.8rem; line-height:1.3;}
.subtitle{color:#667085;font-size:1rem;margin-bottom:1rem;}
.section-title{font-size:1.35rem;font-weight:800;margin-top:.5rem;margin-bottom:.4rem;}
.section-subtitle{color:#667085;font-size:.95rem;margin-bottom:.8rem;}
.route-card{border:1px solid #EAECF0;border-radius:14px;padding:12px;background:#FFF;margin-bottom:10px;}
.small-note{color:#667085;font-size:.85rem;}
.day-card{border:1px solid #D0D5DD;border-radius:16px;padding:16px 18px;background:#F9FAFB;margin:.4rem 0 1rem 0;}
.day-card-title{font-size:.85rem;color:#667085;font-weight:700;text-transform:uppercase;letter-spacing:.04em;margin-bottom:.2rem;}
.day-card-date{font-size:1.45rem;font-weight:850;color:#101828;margin-bottom:.25rem;}
.day-card-detail{font-size:.95rem;color:#475467;}
.metric-card-soft{border:1px solid #EAECF0;border-radius:14px;padding:10px 12px;background:#FFFFFF;}
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-title">Baseline Operacional + Optimizador PM</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle">Comparación automática: Baseline vs Optimización de capacidades</div>', unsafe_allow_html=True)


HUB_LAT = -33.43291
HUB_LON = -70.797027
HUB_COORD = (HUB_LAT, HUB_LON)
TIEMPO_MAX_RUTA_MIN = 480
HORA_CORTE_HUB_MIN = 14 * 60
# Bloques operativos flexibles: no obliga volver exactamente a las 14:00.
# AM debe cerrar cerca del corte; PM puede partir antes/después del corte según ventana.
BLOQUE_AM_START_MIN = 6 * 60
BLOQUE_AM_END_MIN = 15 * 60
BLOQUE_PM_START_MIN = 12 * 60
BLOQUE_PM_END_MIN = 22 * 60
TIEMPO_SERVICIO_BASE = 15
TIEMPO_SERVICIO_OS = 0.1
DEFAULT_CAPACIDAD = 20
APP_DIR = Path(__file__).resolve().parent
DICRUTA_DEFAULT_PATH = APP_DIR / "dicruta.xlsx"
EVENTOS_OPTIMIZABLES = ["PU", "CSC", "CNP", "NHP"]

# Parámetros configurables para producción / Render.
# Se pueden sobrescribir desde Environment Variables sin modificar el código.
ORTOOLS_MAX_PUNTOS = int(os.getenv("ORTOOLS_MAX_PUNTOS", "180"))
ORTOOLS_TIME_LIMIT_SECONDS = int(os.getenv("ORTOOLS_TIME_LIMIT_SECONDS", "20"))
OSRM_MATRIX_TIMEOUT_SECONDS = int(os.getenv("OSRM_MATRIX_TIMEOUT_SECONDS", "20"))
OSRM_ROUTE_TIMEOUT_SECONDS = int(os.getenv("OSRM_ROUTE_TIMEOUT_SECONDS", "12"))
OSRM_MAX_MATRIX_POINTS = int(os.getenv("OSRM_MAX_MATRIX_POINTS", "90"))
CACHE_TTL_SECONDS = int(os.getenv("CACHE_TTL_SECONDS", "3600"))
PENALIZACION_RUTA_NO_PRIORIZADA = int(os.getenv("PENALIZACION_RUTA_NO_PRIORIZADA", "5000"))
# Costo fijo que paga el solver por cada vehículo/vuelta que decide abrir.
# Es la palanca principal para el objetivo "menos rutas, mayor ocupación":
# mientras más alto, más agresivamente el optimizador consolida carga en
# menos vehículos antes de abrir uno adicional. Debe mantenerse menor que
# PENALIZACION_RUTA_NO_PRIORIZADA para que la prioridad de tipo de ruta
# siga pesando más que la sola apertura de un vehículo adicional.
COSTO_FIJO_VEHICULO_ABIERTO = int(os.getenv("COSTO_FIJO_VEHICULO_ABIERTO", "2500"))
DISABLE_OSRM = os.getenv("DISABLE_OSRM", "false").strip().lower() in {"1", "true", "yes", "si", "sí"}



# Capacidad del diccionario viene en m3.
# El volumen operativo del optimizador se calcula en m3 desde kilo mayor.
# Conversión usada: m3 = kilo_mayor / 250, porque kilo_mayor = cm3 / 4000.

# ============================================================
# UTILIDADES
# ============================================================
def normalizar_ruta(valor):
    if pd.isna(valor):
        return np.nan
    return str(valor).upper().strip().replace(".0", "")

def buscar_columna(df, opciones):
    cols = {str(c).lower().strip(): c for c in df.columns}
    for op in opciones:
        if op.lower() in cols:
            return cols[op.lower()]
    return None

def buscar_columna_contiene(df, patrones):
    for c in df.columns:
        cl = str(c).lower().strip()
        if any(p.lower() in cl for p in patrones):
            return c
    return None

def color_from_text(text):
    h = hashlib.md5(str(text).encode()).hexdigest()
    return [int(h[0:2],16), int(h[2:4],16), int(h[4:6],16), 220]

def color_rgba_css(color):
    return f"rgba({color[0]},{color[1]},{color[2]},{color[3]/255:.2f})"

def num(v):
    if pd.isna(v): return "-"
    try: return f"{int(round(float(v))):,}".replace(",", ".")
    except Exception: return str(v)

def dec(v):
    if pd.isna(v): return "-"
    try: return f"{float(v):,.1f}".replace(",", ".")
    except Exception: return str(v)

def kilo_mayor_a_m3(valor):
    """Convierte kilo mayor a m3 para comparar contra capacidad vehicular en m3.

    El kilo mayor viene calculado como ancho * largo * alto / 4000,
    usando dimensiones en cm. Por eso:
        m3 = kilo_mayor * 4000 / 1.000.000 = kilo_mayor / 250
    """
    return (pd.to_numeric(valor, errors="coerce").fillna(0) / 250).clip(lower=0)

def pct(v):
    if pd.isna(v): return "-"
    return f"{float(v):.1f}%"

def minutes_to_time(minutes):
    if pd.isna(minutes): return "-"
    h = int(minutes // 60) % 24
    m = int(minutes % 60)
    return f"{h:02d}:{m:02d}"

def time_to_minutes(t):
    if pd.isna(t): return None
    if hasattr(t, "hour"):
        return int(t.hour) * 60 + int(t.minute)
    try:
        tt = pd.to_datetime(str(t), errors="coerce").time()
        return int(tt.hour) * 60 + int(tt.minute)
    except Exception:
        return None

def extraer_rango_ventana(valor):
    if pd.isna(valor): return (pd.NaT, pd.NaT)
    txt = str(valor).lower().strip()
    txt = txt.replace("hrs", "").replace("hr", "")
    txt = txt.replace("desde", "").replace("hasta", "-")
    txt = txt.replace(" a ", "-").replace("–", "-").replace("—", "-")
    partes = re.findall(r"\d{1,2}(?::\d{2})?", txt)
    if len(partes) < 2: return (pd.NaT, pd.NaT)
    h_ini, h_fin = partes[0], partes[1]
    if ":" not in h_ini: h_ini += ":00"
    if ":" not in h_fin: h_fin += ":00"
    return (pd.to_datetime(h_ini, errors="coerce").time(), pd.to_datetime(h_fin, errors="coerce").time())

def haversine_km(lat1, lon1, lat2, lon2):
    r = 6371
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi, dl = math.radians(lat2-lat1), math.radians(lon2-lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dl/2)**2
    return 2*r*math.atan2(math.sqrt(a), math.sqrt(1-a))


def concatenar_no_vacios(dataframes):
    """Concatena solo DataFrames válidos y evita errores cuando una sección no tiene registros."""
    validos = [df for df in dataframes if isinstance(df, pd.DataFrame) and not df.empty]
    return pd.concat(validos, ignore_index=True) if validos else pd.DataFrame()

# ============================================================
# FILTROS TIPO EXCEL
# ============================================================
def _filter_signature(options):
    return hashlib.md5("|".join([str(o) for o in options]).encode()).hexdigest()[:8]

def _opciones_validas_filtro(options):
    """Ordena y descarta valores nulos: base común de ambos filtros tipo Excel."""
    return sorted([o for o in options if pd.notna(o)])

def _sincronizar_estado_filtro(selected_key, sig_key, sig, valor_inicial):
    """Si el universo de opciones cambió respecto de la última corrida, reinicia
    la selección guardada en session_state. Evita checkboxes/radios pegados de
    un filtro anterior (por ejemplo, al cambiar de fecha o de clase de ruta).
    """
    if sig_key not in st.session_state or st.session_state[sig_key] != sig:
        st.session_state[selected_key] = valor_inicial
        st.session_state[sig_key] = sig

def _filtrar_por_busqueda(options, search):
    return [o for o in options if search.lower() in str(o).lower()] if search else options

def excel_filter(label, options, key, default=None):
    """Filtro tipo Excel (multi-selección) que se resetea cuando cambia el
    universo de opciones. Evita que queden checkboxes antiguos pegados al
    cambiar fecha/clase/tipo.
    """
    options = _opciones_validas_filtro(options)
    if default is None:
        default = options
    default = [o for o in default if o in options]
    sig = _filter_signature(options)
    selected_key = f"{key}_selected"
    sig_key = f"{key}_options_sig"

    _sincronizar_estado_filtro(selected_key, sig_key, sig, list(default))

    # limpia selecciones que ya no existen
    st.session_state[selected_key] = [o for o in st.session_state.get(selected_key, []) if o in options]

    resumen = "Todos" if len(st.session_state[selected_key]) == len(options) else f"{len(st.session_state[selected_key])}/{len(options)}"
    with st.popover(f"🔽 {label}: {resumen}", use_container_width=True):
        search = st.text_input("Buscar", key=f"{key}_{sig}_search")
        visibles = _filtrar_por_busqueda(options, search)
        c1, c2 = st.columns(2)
        if c1.button("Seleccionar todo", key=f"{key}_{sig}_all"):
            st.session_state[selected_key] = list(options)
            st.rerun()
        if c2.button("Limpiar", key=f"{key}_{sig}_clear"):
            st.session_state[selected_key] = []
            st.rerun()
        st.caption(f"Seleccionados: {len(st.session_state[selected_key])} de {len(options)}")
        for op in visibles:
            op_hash = hashlib.md5(str(op).encode()).hexdigest()[:10]
            checked = op in st.session_state[selected_key]
            val = st.checkbox(str(op), value=checked, key=f"{key}_{sig}_chk_{op_hash}")
            if val and op not in st.session_state[selected_key]:
                st.session_state[selected_key].append(op)
            if not val and op in st.session_state[selected_key]:
                st.session_state[selected_key].remove(op)
    return st.session_state[selected_key]

def excel_single_filter(label, options, key, default=None):
    """Selector único que se resetea si cambian las opciones disponibles."""
    options = _opciones_validas_filtro(options)
    if not options:
        return None
    if default is None or default not in options:
        default = options[0]
    sig = _filter_signature(options)
    selected_key = f"{key}_selected_single"
    sig_key = f"{key}_options_sig_single"

    if st.session_state.get(selected_key) not in options:
        st.session_state[sig_key] = None  # fuerza resincronización también si el valor guardado ya no es válido
    _sincronizar_estado_filtro(selected_key, sig_key, sig, default)

    with st.popover(f"🔽 {label}: {st.session_state[selected_key]}", use_container_width=True):
        search = st.text_input("Buscar", key=f"{key}_{sig}_search_single")
        visibles = _filtrar_por_busqueda(options, search)
        if visibles:
            actual = st.session_state[selected_key]
            idx = visibles.index(actual) if actual in visibles else 0
            seleccionado = st.radio("Selecciona", visibles, index=idx, key=f"{key}_{sig}_radio")
            st.session_state[selected_key] = seleccionado
    return st.session_state[selected_key]

# ============================================================
# CARGA Y PREPARACIÓN DE DATA
# ============================================================
@st.cache_data(show_spinner=False, ttl=CACHE_TTL_SECONDS, max_entries=32)
def read_excel_any(file_or_path):
    xl = pd.ExcelFile(file_or_path)
    # usa la hoja con más columnas útiles
    best = None
    best_score = -1
    for sh in xl.sheet_names:
        df = pd.read_excel(file_or_path, sheet_name=sh)
        score = sum(str(c).lower() in ["ruta","reserva","a_operacion","solo_fch","capacidad"] for c in df.columns)
        if score > best_score:
            best, best_score = df, score
    return best

@st.cache_data(show_spinner=False, ttl=CACHE_TTL_SECONDS, max_entries=32)
def asignar_vueltas_por_cierre(peu_df, accesos_df):
    peu_df = peu_df.copy(); accesos_df = accesos_df.copy(); resultados = []
    for (fecha, ruta), g in peu_df.groupby(["fecha_operacion", "ruta"], dropna=False):
        g = g.sort_values("datetime_gestion").copy()
        acc = accesos_df[(accesos_df["solo_fch"] == fecha) & (accesos_df["ruta"] == ruta)].sort_values("datetime_ingreso")
        cierres = acc["datetime_ingreso"].dropna().to_list()
        if not cierres:
            g["vuelta"] = np.nan; g["datetime_cierre_vuelta"] = pd.NaT; g["estado_asignacion_vuelta"] = "Sin ingreso HJ"; resultados.append(g); continue
        cierres_np = np.array(cierres, dtype="datetime64[ns]")
        vueltas=[]; cierre_asig=[]; estados=[]
        for dt in g["datetime_gestion"]:
            idx = np.searchsorted(cierres_np, np.datetime64(dt), side="left")
            vueltas.append(idx+1)
            if idx < len(cierres):
                cierre_asig.append(cierres[idx]); estados.append("Vuelta cerrada por ingreso HJ")
            else:
                cierre_asig.append(pd.NaT); estados.append("Vuelta posterior al último ingreso HJ")
        g["vuelta"] = vueltas; g["datetime_cierre_vuelta"] = cierre_asig; g["estado_asignacion_vuelta"] = estados
        resultados.append(g)
    return pd.concat(resultados, ignore_index=True) if resultados else peu_df

# ============================================================
# OPTIMIZADOR
# ============================================================
def fallback_matrix(coords, speed_kmh=28):
    n = len(coords); matrix = [[0]*n for _ in range(n)]
    for i,a in enumerate(coords):
        for j,b in enumerate(coords):
            if i == j: continue
            km = haversine_km(a[0], a[1], b[0], b[1])
            matrix[i][j] = int(math.ceil((km*1.35)/speed_kmh*60))
    return matrix

@st.cache_data(show_spinner=False, ttl=CACHE_TTL_SECONDS, max_entries=32)
def get_vial_matrix(coords_tuple, usar_osrm=True):
    coords = list(coords_tuple)
    if DISABLE_OSRM or not usar_osrm or len(coords) > OSRM_MAX_MATRIX_POINTS:
        return fallback_matrix(coords), "fallback"
    coord_txt = ";".join([f"{lon},{lat}" for lat, lon in coords])
    url = f"https://router.project-osrm.org/table/v1/driving/{coord_txt}?annotations=duration"
    try:
        r = requests.get(url, timeout=OSRM_MATRIX_TIMEOUT_SECONDS)
        data = r.json()
        if data.get("durations"):
            matriz = np.array(data["durations"], dtype=float) / 60.0
            return np.ceil(np.nan_to_num(matriz, nan=9999)).astype(int).tolist(), "osrm"
    except Exception:
        pass
    return fallback_matrix(coords), "fallback"

@st.cache_data(show_spinner=False, ttl=CACHE_TTL_SECONDS, max_entries=32)
def get_route_geometry(coords_tuple, usar_osrm=True):
    coords = list(coords_tuple)
    if DISABLE_OSRM or not usar_osrm or len(coords) < 2:
        return [[lat, lon] for lat, lon in coords], False, sum(haversine_km(a[0],a[1],b[0],b[1])*1.35 for a,b in zip(coords[:-1], coords[1:]))
    coords_use = coords
    if len(coords_use) > 25:
        idx = np.linspace(0, len(coords_use)-1, 25).astype(int)
        coords_use = [coords_use[i] for i in idx]
    coord_txt = ";".join([f"{lon},{lat}" for lat, lon in coords_use])
    url = f"https://router.project-osrm.org/route/v1/driving/{coord_txt}"
    try:
        r = requests.get(url, params={"overview":"full", "geometries":"geojson", "steps":"false"}, timeout=OSRM_ROUTE_TIMEOUT_SECONDS)
        routes = r.json().get("routes", [])
        if routes:
            geom = [[p[1], p[0]] for p in routes[0]["geometry"]["coordinates"]]
            return geom, True, routes[0].get("distance", 0)/1000
    except Exception:
        pass
    return [[lat, lon] for lat, lon in coords], False, sum(haversine_km(a[0],a[1],b[0],b[1])*1.35 for a,b in zip(coords[:-1], coords[1:]))


def construir_flotas_capacidades(flota_baseline, demanda_total=0, tipos_ruta_priorizados=None):
    """
    Optimización de capacidades.

    - Usa el mismo universo de vehículos/rutas del baseline del día y permite usar menos.
    - Respeta todas las restricciones operacionales.
    - Cuando se seleccionan tipos de ruta prioritarios, intenta asignarles carga primero.
    - La prioridad es suave: si esas rutas no son suficientes o compatibles, usa el resto.
    """
    tipos_ruta_priorizados = {
        str(x).strip().upper()
        for x in (tipos_ruta_priorizados or [])
    }

    flota = flota_baseline.copy()
    if flota.empty:
        flota = pd.DataFrame([
            {
                "vehiculo_base": f"CAP-{i+1}",
                "capacidad": DEFAULT_CAPACIDAD,
                "tipo_ruta": "SIN INFORMACIÓN"
            }
            for i in range(max(1, math.ceil(demanda_total/DEFAULT_CAPACIDAD)))
        ])

    flota["capacidad"] = pd.to_numeric(
        flota["capacidad"],
        errors="coerce"
    ).fillna(DEFAULT_CAPACIDAD)

    if "tipo_ruta" not in flota.columns:
        flota["tipo_ruta"] = "SIN INFORMACIÓN"

    flota["tipo_ruta_norm"] = (
        flota["tipo_ruta"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    flota["es_tipo_priorizado"] = (
        flota["tipo_ruta_norm"].isin(tipos_ruta_priorizados)
        if tipos_ruta_priorizados
        else False
    )

    return flota.sort_values(
        ["es_tipo_priorizado", "capacidad", "vehiculo_base"],
        ascending=[False, False, True]
    ).reset_index(drop=True)

# El motor de optimización se define en la sección V18 antes de la interfaz.

# ============================================================
# MÉTRICAS Y MAPAS
# ============================================================

def metricas_opt(resumen, puntos, escenario, rutas_base_objetivo=None):
    """
    Métricas de escenarios optimizados.

    Reglas corregidas v7:
    - Optimización de ruta mantiene la misma flota física del baseline. Por eso
      cantidad_rutas y productividad se calculan con rutas_base_objetivo.
    - Las reservas no asignadas se excluyen del resumen y se informan en una tabla separada.
      optimizadas, vueltas, capacidad ni ocupación.
    - factor_ocupacion se expresa como porcentaje real: volumen_m3 / capacidad_m3 * 100.
    """
    total_os = puntos["os"].sum() if not puntos.empty else 0
    reservas = puntos["id_punto_opt"].nunique() if not puntos.empty else 0
    volumen = puntos["volumen"].sum() if not puntos.empty else 0

    if resumen.empty:
        rutas_calc = rutas_base_objetivo if rutas_base_objetivo is not None else 0
        return {"escenario":escenario, "cantidad_rutas":rutas_calc, "vueltas":0, "total_os":total_os,
                "reservas":reservas, "volumen":volumen, "capacidad_total":np.nan, "cumplimiento_vh":np.nan, "excepciones":np.nan,
                "productividad": total_os/rutas_calc if rutas_calc else np.nan, "factor_ocupacion":np.nan, "tiempo_ruta":np.nan}

    resumen_ok = resumen.copy()
    resumen_diag = pd.DataFrame()

    if resumen_ok.empty:
        rutas_calc = rutas_base_objetivo if rutas_base_objetivo is not None else 0
        return {"escenario":escenario, "cantidad_rutas":rutas_calc, "vueltas":0, "total_os":total_os,
                "reservas":reservas, "volumen":volumen, "capacidad_total":np.nan, "cumplimiento_vh":0 if not resumen_diag.empty else np.nan, "excepciones":np.nan,
                "productividad": total_os/rutas_calc if rutas_calc else np.nan, "factor_ocupacion":np.nan, "tiempo_ruta":np.nan}

    if "activo" not in resumen_ok.columns:
        resumen_ok["activo"] = pd.to_numeric(resumen_ok.get("paradas", 0), errors="coerce").fillna(0) > 0
    else:
        resumen_ok["activo"] = resumen_ok["activo"].fillna(True).astype(bool)

    resumen_activo = resumen_ok[resumen_ok["activo"]].copy()
    rutas_usadas = resumen_activo["vehiculo_base"].nunique()
    rutas_fisicas = int(rutas_base_objetivo) if rutas_base_objetivo is not None else int(rutas_usadas)

    # v10: ocupación macro = promedio de ocupación vehículo-día.
    # En Optimización de ruta se incluyen vehículos sin carga con ocupación 0,
    # porque la flota se mantiene. En Optimización de capacidades solo quedan
    # vehículos efectivamente activados.
    occ_veh_por_dia = resumen_ok.groupby("vehiculo_base")["factor_ocupacion"].mean() if "factor_ocupacion" in resumen_ok.columns else pd.Series(dtype=float)
    occ_veh_dia = occ_veh_por_dia.mean() if len(occ_veh_por_dia) else np.nan
    factor_macro = occ_veh_dia
    capacidad_total = resumen_activo["capacidad"].sum() if "capacidad" in resumen_activo.columns and not resumen_activo.empty else np.nan

    return {
        "escenario":escenario,
        "cantidad_rutas":rutas_fisicas,
        "vueltas":len(resumen_activo),
        "total_os":total_os,
        "reservas":reservas,
        "volumen":volumen,
        "capacidad_total":capacidad_total,
        "cumplimiento_vh":100.0 if resumen_diag.empty else np.nan,
        "excepciones":np.nan,
        "productividad": total_os/rutas_fisicas if rutas_fisicas else np.nan,
        "factor_ocupacion": factor_macro,
        "ocupacion_vehiculo_dia": occ_veh_dia,
        "tiempo_ruta": resumen_activo["tiempo_ruta_min"].mean() if not resumen_activo.empty else 0
    }

def resumen_por_ruta_baseline(baseline, dicruta, cols):
    col_reserva, col_q_os, col_km = cols["col_reserva"], cols["col_q_os"], cols["col_km"]
    cap_map = dicruta.set_index("ruta")["capacidad"].to_dict()
    d = baseline.copy()
    if col_km:
        q_os_base = pd.to_numeric(d[col_q_os], errors="coerce").fillna(1) if col_q_os else pd.Series(1, index=d.index)
        ajuste_vol = ajustar_volumen_m3(d[col_km], q_os_base, ajuste_vol_por_os=d["ruta"].map(dicruta.set_index("ruta")["ajuste_vol"].to_dict()) if "ajuste_vol" in dicruta.columns else None, umbral_m3_por_os=3.0, m3_estandar_por_os=0.33)
        d["volumen_original_m3"] = ajuste_vol["volumen_original_m3"]
        d["volumen"] = ajuste_vol["volumen_m3"]
        d["m3_por_os"] = ajuste_vol["m3_por_os"]
        d["volumen_ajustado"] = ajuste_vol["volumen_ajustado"]
    else:
        d["volumen_original_m3"] = 0
        d["volumen"] = 0
        d["m3_por_os"] = 0
        d["volumen_ajustado"] = False
    d["capacidad_ruta"] = d["ruta"].map(cap_map).fillna(DEFAULT_CAPACIDAD)
    occ_vuelta_tmp = d.groupby(["ruta","vuelta"], dropna=False).agg(
        volumen_vuelta=("volumen", "sum"),
        capacidad=("capacidad_ruta", "first")
    ).reset_index()
    occ_vuelta_tmp["factor_ocupacion_vuelta"] = np.where(
        occ_vuelta_tmp["capacidad"] > 0,
        occ_vuelta_tmp["volumen_vuelta"] / occ_vuelta_tmp["capacidad"] * 100,
        np.nan
    )
    occ_ruta_map = occ_vuelta_tmp.groupby("ruta")["factor_ocupacion_vuelta"].mean().to_dict()
    out = d.groupby(["fecha_operacion","ruta"], dropna=False).agg(
        vueltas=("vuelta", lambda s: s.dropna().nunique()), reservas=(col_reserva,"nunique"), total_os=(col_q_os,"sum"),
        volumen=("volumen","sum"), volumen_original_m3=("volumen_original_m3", "sum"), reservas_ajustadas=("volumen_ajustado", "sum"),
        cumplimiento_vh=("cumple_ventana","mean"), excepciones=("es_excepcion","mean"),
        primer_pu=("datetime_gestion","min"), ultimo_pu=("datetime_gestion","max")
    ).reset_index()
    out["capacidad"] = out["ruta"].map(cap_map).fillna(DEFAULT_CAPACIDAD)
    out["factor_ocupacion"] = out["ruta"].map(occ_ruta_map)
    out["productividad"] = out["total_os"] / 1
    out["cumplimiento_vh"] *= 100; out["excepciones"] *= 100
    out["escenario"] = "Baseline"
    return out

def resumen_por_vuelta_baseline(baseline, dicruta, cols):
    col_reserva, col_q_os, col_km = cols["col_reserva"], cols["col_q_os"], cols["col_km"]
    cap_map = dicruta.set_index("ruta")["capacidad"].to_dict()
    d = baseline.copy()
    if col_km:
        q_os_base = pd.to_numeric(d[col_q_os], errors="coerce").fillna(1) if col_q_os else pd.Series(1, index=d.index)
        ajuste_vol = ajustar_volumen_m3(d[col_km], q_os_base, ajuste_vol_por_os=d["ruta"].map(dicruta.set_index("ruta")["ajuste_vol"].to_dict()) if "ajuste_vol" in dicruta.columns else None, umbral_m3_por_os=3.0, m3_estandar_por_os=0.33)
        d["volumen_original_m3"] = ajuste_vol["volumen_original_m3"]
        d["volumen"] = ajuste_vol["volumen_m3"]
        d["m3_por_os"] = ajuste_vol["m3_por_os"]
        d["volumen_ajustado"] = ajuste_vol["volumen_ajustado"]
    else:
        d["volumen_original_m3"] = 0
        d["volumen"] = 0
        d["m3_por_os"] = 0
        d["volumen_ajustado"] = False
    out = d.groupby(["fecha_operacion","ruta","vuelta"], dropna=False).agg(
        reservas=(col_reserva,"nunique"), total_os=(col_q_os,"sum"), volumen=("volumen","sum"),
        volumen_original_m3=("volumen_original_m3", "sum"), reservas_ajustadas=("volumen_ajustado", "sum"),
        cumplimiento_vh=("cumple_ventana","mean"), excepciones=("es_excepcion","mean"),
        primer_pu=("datetime_gestion","min"), ultimo_pu=("datetime_gestion","max"), cierre_vuelta=("datetime_cierre_vuelta","max")
    ).reset_index()
    out["capacidad"] = out["ruta"].map(cap_map).fillna(DEFAULT_CAPACIDAD)
    out["factor_ocupacion"] = out["volumen"] / out["capacidad"] * 100
    out["cumplimiento_vh"] *= 100; out["excepciones"] *= 100
    out["escenario"] = "Baseline"
    return out

def preparar_mapa_baseline(baseline, cols):
    lat_col, lon_col = cols["lat_col"], cols["lon_col"]
    if not lat_col or not lon_col:
        return pd.DataFrame(), pd.DataFrame()

    mapa = baseline.dropna(subset=[lat_col, lon_col]).sort_values(
        ["ruta", "vuelta", "datetime_gestion"]
    ).copy()

    if mapa.empty:
        return pd.DataFrame(), pd.DataFrame()

    mapa["orden"] = mapa.groupby(["ruta", "vuelta"]).cumcount() + 1
    mapa["orden_txt"] = mapa["orden"].astype(str)

    # Bloque por hora real de gestión
    mapa["bloque"] = np.where(
        mapa["datetime_gestion"].dt.hour < 14,
        "AM",
        "PM"
    )

    mapa["color"] = mapa.apply(
        lambda r: color_from_text(f"BL-{r['ruta']}-{r['vuelta']}"),
        axis=1
    )

    paths = []

    # Se arma una ruta por ruta + vuelta + bloque
    for (ruta, vuelta, bloque), g in mapa.groupby(["ruta", "vuelta", "bloque"], dropna=False):
        coords = [HUB_COORD] + [
            (float(r[lat_col]), float(r[lon_col])) for _, r in g.iterrows()
        ] + [HUB_COORD]

        geom, ok, km = get_route_geometry(tuple(coords), usar_osrm=True)

        paths.append({
            "ruta": ruta,
            "vehiculo_base": ruta,
            "vuelta": str(vuelta),
            "bloque": bloque,
            "id": f"{ruta}-{vuelta}-{bloque}",
            "path": [[lon, lat] for lat, lon in geom],
            "color": color_from_text(f"BL-{ruta}-{vuelta}-{bloque}"),
            "km_estimado": km
        })

    mapa = mapa.rename(columns={lat_col: "lat", lon_col: "lon"})

    return mapa, pd.DataFrame(paths)

def preparar_mapa_opt(routes):
    stops=[]; paths=[]
    for r in routes:
        veh = str(r.get("vehiculo_base", ""))
        vuelta = str(r.get("vuelta", ""))
        paths.append({
                "id": f"{veh}-V{vuelta}",
                "vehiculo_base": veh,
                "ruta": veh,
                "vuelta": vuelta,
                "bloque": r.get("bloque", ""),
                "path": [[lon,lat] for lat,lon in r["geometry"]],
                "color": r["color"]
        })
        for s in r["stops"]:
            row=s.copy(); row["color"]=r["color"]; row["orden_txt"]=str(row["secuencia"]); stops.append(row)
    return pd.DataFrame(stops), pd.DataFrame(paths)

def render_map(stops, paths, title, key_prefix="mapa"):
    st.markdown(f"#### {title}")

    if stops.empty or paths.empty:
        st.warning("No hay puntos para mostrar.")
        return

    # Filtro por bloque AM / PM
    if "bloque" in stops.columns:
        bloques_disponibles = sorted([str(x) for x in stops["bloque"].dropna().unique()])

        bloque_sel = st.selectbox(
            "Ver bloque",
            ["Todos"] + bloques_disponibles,
            key=f"{key_prefix}_bloque_visible"
        )

        if bloque_sel != "Todos":
            stops = stops[stops["bloque"].astype(str) == bloque_sel].copy()

            if "bloque" in paths.columns:
                paths = paths[paths["bloque"].astype(str) == bloque_sel].copy()

    # Filtro por ruta / vehículo
    ruta_col = "vehiculo_base" if "vehiculo_base" in stops.columns else "ruta"
    opciones = sorted([str(x) for x in stops[ruta_col].dropna().unique()])

    seleccion = st.selectbox(
        "Ver ruta / vehículo",
        ["Todas"] + opciones,
        key=f"{key_prefix}_ruta_visible"
    )

    if seleccion != "Todas":
        stops = stops[stops[ruta_col].astype(str) == seleccion].copy()

        if "vehiculo_base" in paths.columns:
            paths = paths[paths["vehiculo_base"].astype(str) == seleccion].copy()
        elif "ruta" in paths.columns:
            paths = paths[paths["ruta"].astype(str) == seleccion].copy()
        else:
            paths = paths[paths["id"].astype(str).str.startswith(seleccion)].copy()

    if stops.empty or paths.empty:
        st.warning("No hay puntos para la selección realizada.")
        return

    n_rutas_visibles = stops[ruta_col].nunique() if ruta_col in stops.columns else paths["id"].nunique()
    n_paradas_visibles = len(stops)
    os_visibles = pd.to_numeric(stops["os"], errors="coerce").sum() if "os" in stops.columns else None
    vol_visible = pd.to_numeric(stops["volumen"], errors="coerce").sum() if "volumen" in stops.columns else None

    stat_txt = f"🛣️ {num(n_rutas_visibles)} ruta(s)/vehículo(s) · 📍 {num(n_paradas_visibles)} parada(s)"
    if os_visibles is not None:
        stat_txt += f" · 📦 {num(os_visibles)} OS"
    if vol_visible is not None:
        stat_txt += f" · 🧊 {dec(vol_visible)} m³"
    st.caption(stat_txt)

    path_layer = pdk.Layer(
        "PathLayer",
        data=paths,
        get_path="path",
        get_width=8,
        width_min_pixels=4,
        get_color="color",
        pickable=True
    )

    point_layer = pdk.Layer(
        "ScatterplotLayer",
        data=stops,
        get_position="[lon, lat]",
        get_radius=105,
        get_fill_color="color",
        get_line_color=[255, 255, 255],
        line_width_min_pixels=3,
        pickable=True
    )

    text_layer = pdk.Layer(
        "TextLayer",
        data=stops,
        get_position="[lon, lat]",
        get_text="orden_txt",
        get_size=14,
        get_color=[255, 255, 255, 255],
        get_text_anchor="'middle'",
        get_alignment_baseline="'center'"
    )

    hub = pd.DataFrame([{
        "lat": HUB_LAT,
        "lon": HUB_LON,
        "color": [255, 170, 0, 245],
        "nombre": "HUB1"
    }])

    hub_layer = pdk.Layer(
        "ScatterplotLayer",
        data=hub,
        get_position="[lon, lat]",
        get_radius=220,
        get_fill_color="color",
        get_line_color=[255, 255, 255],
        line_width_min_pixels=3,
        pickable=True
    )

    view = pdk.ViewState(
        latitude=stops["lat"].mean(),
        longitude=stops["lon"].mean(),
        zoom=10,
        pitch=0
    )

    tooltip = {
        "html": "<b>ID:</b> {id_punto_opt}<br/><b>Ruta:</b> {ruta}<br/><b>Vehículo:</b> {vehiculo_base}<br/><b>Vuelta:</b> {vuelta}<br/><b>Bloque:</b> {bloque}<br/><b>Orden:</b> {orden_txt}<br/><b>ETA:</b> {eta}<br/><b>OS:</b> {os}<br/><b>Volumen m³:</b> {volumen}",
        "style": {"backgroundColor": "white", "color": "black"}
    }

    st.pydeck_chart(
        pdk.Deck(
            map_style="light",
            layers=[path_layer, point_layer, text_layer, hub_layer],
            initial_view_state=view,
            tooltip=tooltip
        ),
        use_container_width=True
    )

# ============================================================
# MOTOR V18 · RESTRICCIONES OPERACIONALES CENTRALIZADAS
# ============================================================
# Estas redefiniciones reemplazan el motor anterior sin eliminar la interfaz,
# mapas, filtros, gráficos ni exportaciones existentes.

DICRUTA_VALIDACIONES = []


def _normalizar_nombre_columna(nombre):
    txt = str(nombre).strip().lower()
    txt = txt.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
    txt = re.sub(r"[^a-z0-9]+", "_", txt).strip("_")
    return txt


def _a_numero_nullable(serie):
    return pd.to_numeric(serie, errors="coerce")


def _a_booleano_propio(valor):
    if pd.isna(valor):
        return False
    txt = str(valor).strip().upper().replace("Í", "I")
    return txt in {"SI", "S", "1", "TRUE", "YES", "Y", "VERDADERO"}


def _excel_hora_a_minutos(valor):
    """Admite time/datetime, texto HH:MM, timestamp y fracción horaria Excel."""
    if pd.isna(valor) or str(valor).strip() == "":
        return np.nan
    if isinstance(valor, (int, float, np.integer, np.floating)):
        v = float(valor)
        if 0 <= v < 1:
            return int(round(v * 24 * 60)) % 1440
        if 0 <= v <= 24:
            return int(round(v * 60)) % 1440
        # serial Excel con fecha + fracción
        frac = v % 1
        if frac > 0:
            return int(round(frac * 24 * 60)) % 1440
    if hasattr(valor, "hour"):
        return int(valor.hour) * 60 + int(getattr(valor, "minute", 0))
    txt = str(valor).strip()
    m = re.search(r"(?:^|\s)(\d{1,2}):(\d{2})(?::\d{2})?", txt)
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        return h * 60 + mi if h < 24 and mi < 60 else np.nan
    dt = pd.to_datetime(txt, errors="coerce")
    if pd.notna(dt):
        return int(dt.hour) * 60 + int(dt.minute)
    return np.nan


def _format_rango(inicio, fin):
    if pd.isna(inicio) and pd.isna(fin): return "Sin restricción"
    if pd.isna(fin): return f"Desde {minutes_to_time(inicio)}"
    if pd.isna(inicio): return f"Hasta {minutes_to_time(fin)}"
    return f"{minutes_to_time(inicio)}–{minutes_to_time(fin)}"


@st.cache_data(show_spinner=False, ttl=CACHE_TTL_SECONDS, max_entries=32)
def cargar_dicruta(file_or_path):
    global DICRUTA_VALIDACIONES
    dic = read_excel_any(file_or_path).copy()
    dic.columns = [str(c).strip() for c in dic.columns]
    normalizadas = {_normalizar_nombre_columna(c): c for c in dic.columns}

    aliases = {
        "ruta": ["ruta"], "base": ["base"], "tipo": ["tipo"],
        "tipo_ruta": ["tipo_de_ruta", "tipo_ruta"], "tipo_flota": ["tipo_flota"],
        "capacidad": ["capacidad"], "hora_min": ["hora_min"], "hora_max": ["hora_max"],
        "vol_min": ["vol_min"], "vol_max": ["vol_max"], "resv_max": ["resv_max"],
        "bt": ["bt"], "ajuste_vol": ["ajustevol", "ajuste_vol"], "propio": ["propio"]
    }
    def col(canon):
        for a in aliases[canon]:
            if a in normalizadas: return normalizadas[a]
        return None

    if col("ruta") is None or col("capacidad") is None:
        raise ValueError("El diccionario debe contener las columnas Ruta y Capacidad.")

    out = pd.DataFrame(index=dic.index)
    out["ruta"] = dic[col("ruta")].apply(normalizar_ruta)
    for c in ["base", "tipo", "tipo_ruta", "tipo_flota"]:
        out[c] = dic[col(c)].astype("string").fillna("SIN INFORMACIÓN").str.strip() if col(c) else "SIN INFORMACIÓN"
    out["capacidad"] = _a_numero_nullable(dic[col("capacidad")])
    out["hora_min_minutos"] = dic[col("hora_min")].apply(_excel_hora_a_minutos) if col("hora_min") else np.nan
    out["hora_max_minutos"] = dic[col("hora_max")].apply(_excel_hora_a_minutos) if col("hora_max") else np.nan
    out["hora_min"] = out["hora_min_minutos"].apply(lambda x: minutes_to_time(x) if pd.notna(x) else np.nan)
    out["hora_max"] = out["hora_max_minutos"].apply(lambda x: minutes_to_time(x) if pd.notna(x) else np.nan)
    out["horario_min"] = out["hora_min_minutos"]  # compatibilidad con la interfaz previa
    out["horario_txt"] = out.apply(lambda r: _format_rango(r["hora_min_minutos"], r["hora_max_minutos"]), axis=1)
    for c in ["vol_min", "vol_max", "resv_max", "bt", "ajuste_vol"]:
        out[c] = _a_numero_nullable(dic[col(c)]) if col(c) else np.nan
    out["es_propio"] = dic[col("propio")].apply(_a_booleano_propio) if col("propio") else False

    validaciones = []
    dup = sorted(out.loc[out["ruta"].duplicated(keep=False), "ruta"].dropna().unique().tolist())
    if dup: validaciones.append({"tipo":"Duplicado", "rutas":", ".join(dup), "detalle":"La ruta aparece más de una vez; se utilizará la primera configuración y se informa el duplicado."})
    for i, r in out.iterrows():
        ruta = r["ruta"]
        if pd.isna(ruta): continue
        if pd.isna(r["capacidad"]) or r["capacidad"] <= 0:
            validaciones.append({"tipo":"Capacidad inválida", "rutas":ruta, "detalle":"Capacidad debe ser mayor que cero."})
        if pd.notna(r["vol_min"]) and r["vol_min"] < 0:
            validaciones.append({"tipo":"Vol-min inválido", "rutas":ruta, "detalle":"Vol-min no puede ser negativo."})
        if pd.notna(r["vol_max"]) and r["vol_max"] < 0:
            validaciones.append({"tipo":"Vol-max inválido", "rutas":ruta, "detalle":"Vol-max no puede ser negativo."})
        if pd.notna(r["vol_min"]) and pd.notna(r["vol_max"]) and r["vol_min"] > r["vol_max"]:
            validaciones.append({"tipo":"Rango de volumen inválido", "rutas":ruta, "detalle":"Vol-min es mayor que Vol-max."})
        if pd.notna(r["hora_min_minutos"]) and pd.notna(r["hora_max_minutos"]) and r["hora_min_minutos"] > r["hora_max_minutos"]:
            validaciones.append({"tipo":"Horario inválido", "rutas":ruta, "detalle":"Hora-min es posterior a Hora-max."})
        if pd.notna(r["resv_max"]) and (r["resv_max"] <= 0 or abs(r["resv_max"] - round(r["resv_max"])) > 1e-9):
            validaciones.append({"tipo":"Resv-max inválido", "rutas":ruta, "detalle":"Resv-max debe ser entero positivo."})
        if pd.notna(r["bt"]) and (r["bt"] < 0 or abs(r["bt"] - round(r["bt"])) > 1e-9):
            validaciones.append({"tipo":"BT inválido", "rutas":ruta, "detalle":"BT debe ser entero mayor o igual a cero."})
        if pd.notna(r["ajuste_vol"]) and r["ajuste_vol"] <= 0:
            validaciones.append({"tipo":"AjusteVol inválido", "rutas":ruta, "detalle":"Se usará respaldo de 0,33 m³ por OS."})
    DICRUTA_VALIDACIONES = validaciones
    out["config_valida"] = True
    rutas_invalidas = {v["rutas"] for v in validaciones if "," not in v["rutas"] and v["tipo"] != "AjusteVol inválido"}
    out.loc[out["ruta"].isin(rutas_invalidas), "config_valida"] = False
    out["capacidad"] = out["capacidad"].where(out["capacidad"] > 0, np.nan)
    out["resv_max"] = out["resv_max"].where((out["resv_max"] > 0) & (out["resv_max"] % 1 == 0), np.nan)
    out["bt"] = out["bt"].where((out["bt"] >= 0) & (out["bt"] % 1 == 0), np.nan)
    out["ajuste_vol"] = out["ajuste_vol"].where(out["ajuste_vol"] > 0, np.nan)
    return out.dropna(subset=["ruta"]).drop_duplicates("ruta", keep="first").reset_index(drop=True)


@st.cache_data(show_spinner=False, ttl=CACHE_TTL_SECONDS, max_entries=32)
def preparar_data(peu_raw, accesos_raw):
    peu, accesos = peu_raw.copy(), accesos_raw.copy()
    peu.columns = peu.columns.str.strip(); accesos.columns = accesos.columns.str.strip()
    col_ruta_peu = buscar_columna(peu, ["ruta", "tab_ruta", "bdga_cdg"])
    col_ruta_acc = buscar_columna(accesos, ["ruta"])
    col_clase = buscar_columna(peu, ["clase_ruta", "clase de ruta", "base2"])
    col_tipo = buscar_columna(peu, ["base1", "base 1", "tipo_ruta", "tipo de ruta", "operacion"])
    col_fecha_peu = buscar_columna(peu, ["a_operacion", "fecha_operacion", "fecha de operación"])
    col_hora_peu = buscar_columna(peu, ["hora_gestion", "hora de gestion", "hora de gestión"])
    col_fecha_acc = buscar_columna(accesos, ["solo_fch", "solo fecha", "fecha_ingreso", "fecha de ingreso"])
    col_hora_acc = buscar_columna(accesos, ["hora_entrada", "hora entrada", "hora de entrada"])
    col_evento = buscar_columna(peu, ["evento"])
    col_ventana = buscar_columna(peu, ["ventana_horaria", "ventana horaria", "ventana"])
    col_inicio_vh = buscar_columna(peu, ["inicio_vh", "inicio ventana", "ventana_inicio"])
    col_fin_vh = buscar_columna(peu, ["termino_vh", "fin_vh", "fin ventana", "ventana_fin"])
    col_reserva = buscar_columna(peu, ["reserva"])
    col_comuna = buscar_columna(peu, ["comuna", "comuna_destino", "comuna destino"])
    col_q_os = buscar_columna(peu, ["q_os", "cantidad_os", "cantidad de os"])
    col_q_os_b = buscar_columna(peu, ["q_os_b", "q os b", "big ticket"])
    col_km = buscar_columna(peu, ["kilomayor", "kilo_mayor", "kilo mayor", "kmayor", "kmay_ajustado"])
    col_piezas = buscar_columna(peu, ["q_piezas", "piezas", "cantidad_piezas", "volumen"])
    faltantes=[]
    for nombre,c in {"ruta PU":col_ruta_peu,"ruta accesos":col_ruta_acc,"clase/base2 PU":col_clase,"fecha PU":col_fecha_peu,"hora PU":col_hora_peu,"fecha accesos":col_fecha_acc,"hora accesos":col_hora_acc,"evento PU":col_evento,"reserva PU":col_reserva}.items():
        if c is None: faltantes.append(nombre)
    if faltantes: return None,None,None,faltantes
    peu["ruta"] = peu[col_ruta_peu].apply(normalizar_ruta); accesos["ruta"] = accesos[col_ruta_acc].apply(normalizar_ruta)
    peu["clase_ruta"] = peu[col_clase].astype(str).str.upper().str.strip()
    peu["tipo_ruta"] = peu[col_tipo].astype(str).str.upper().str.strip() if col_tipo else "SIN TIPO"
    peu["comuna"] = peu[col_comuna].astype(str).str.upper().str.strip() if col_comuna else "SIN COMUNA"
    peu["fecha_operacion"] = pd.to_datetime(peu[col_fecha_peu], errors="coerce").dt.date
    accesos["solo_fch"] = pd.to_datetime(accesos[col_fecha_acc], errors="coerce").dt.date
    peu["datetime_gestion"] = pd.to_datetime(peu["fecha_operacion"].astype(str)+" "+peu[col_hora_peu].astype(str), errors="coerce")
    accesos["datetime_ingreso"] = pd.to_datetime(accesos["solo_fch"].astype(str)+" "+accesos[col_hora_acc].astype(str), errors="coerce")
    lat_col=buscar_columna_contiene(peu,["lat"]); lon_col=buscar_columna_contiene(peu,["lon"])
    for c in [lat_col,lon_col,col_q_os,col_q_os_b,col_km,col_piezas]:
        if c: peu[c]=pd.to_numeric(peu[c],errors="coerce")
    for c in [col_q_os,col_q_os_b,col_km,col_piezas]:
        if c: peu[c]=peu[c].fillna(0)
    peu["evento_norm"] = peu[col_evento].astype(str).str.upper().str.strip()
    peu["gestion_correcta"] = peu["evento_norm"].eq("PU"); peu["es_excepcion"] = ~peu["gestion_correcta"]
    peu["sin_geo"] = peu[lat_col].isna()|peu[lon_col].isna() if lat_col and lon_col else True
    if col_inicio_vh and col_fin_vh:
        peu["ventana_inicio"] = peu[col_inicio_vh].apply(lambda x: pd.NaT if pd.isna(x) else pd.to_datetime(str(x),errors="coerce").time())
        peu["ventana_fin"] = peu[col_fin_vh].apply(lambda x: pd.NaT if pd.isna(x) else pd.to_datetime(str(x),errors="coerce").time())
    elif col_ventana:
        rangos=peu[col_ventana].apply(extraer_rango_ventana); peu["ventana_inicio"]=rangos.apply(lambda x:x[0]); peu["ventana_fin"]=rangos.apply(lambda x:x[1])
    else: peu["ventana_inicio"]=pd.NaT; peu["ventana_fin"]=pd.NaT
    def cumple_vh(row):
        if pd.isna(row["ventana_inicio"]) or pd.isna(row["ventana_fin"]) or pd.isna(row["datetime_gestion"]): return np.nan
        fecha=row["datetime_gestion"].date(); ini=pd.Timestamp.combine(fecha,row["ventana_inicio"]); fin=pd.Timestamp.combine(fecha,row["ventana_fin"])
        if row["ventana_inicio"]>row["ventana_fin"]: fin+=pd.Timedelta(days=1)
        return ini<=row["datetime_gestion"]<=fin
    peu["cumple_ventana"]=peu.apply(cumple_vh,axis=1)
    peu["estado_ventana"]=np.where(peu["cumple_ventana"].eq(True),"Cumple",np.where(peu["cumple_ventana"].eq(False),"No cumple","Sin ventana válida"))
    peu=peu.dropna(subset=["fecha_operacion","datetime_gestion","ruta"]); accesos=accesos.dropna(subset=["solo_fch","datetime_ingreso","ruta"])
    cols=dict(col_reserva=col_reserva,col_comuna=col_comuna,col_q_os=col_q_os,col_q_os_b=col_q_os_b,col_km=col_km,col_piezas=col_piezas,lat_col=lat_col,lon_col=lon_col)
    return peu,accesos,cols,[]


def ajustar_volumen_m3(kilo_mayor, q_os, ajuste_vol_por_os=None, umbral_m3_por_os=3.0, m3_estandar_por_os=0.33):
    volumen_original=kilo_mayor_a_m3(kilo_mayor)
    os=pd.to_numeric(q_os,errors="coerce").fillna(1).mask(lambda s:s<=0,1)
    if ajuste_vol_por_os is None: ajuste=pd.Series(m3_estandar_por_os,index=volumen_original.index)
    else:
        ajuste=pd.to_numeric(ajuste_vol_por_os,errors="coerce")
        if not isinstance(ajuste,pd.Series): ajuste=pd.Series(ajuste,index=volumen_original.index)
        ajuste=ajuste.reindex(volumen_original.index).fillna(m3_estandar_por_os).mask(lambda s:s<=0,m3_estandar_por_os)
    m3_os=(volumen_original/os).replace([np.inf,-np.inf],0).fillna(0)
    flag=m3_os>umbral_m3_por_os
    final=volumen_original.mask(flag,os*ajuste).clip(lower=0)
    return pd.DataFrame({"volumen_original_m3":volumen_original,"volumen_m3":final,"m3_por_os":m3_os,"ajuste_vol_usado":ajuste,"volumen_ajustado":flag})


def preparar_puntos(df, cols, max_capacidad=None, dicruta=None):
    lat_col,lon_col=cols["lat_col"],cols["lon_col"]; col_reserva=cols["col_reserva"]; col_q_os=cols["col_q_os"]; col_q_os_b=cols.get("col_q_os_b"); col_km=cols["col_km"]
    if not lat_col or not lon_col: return pd.DataFrame()
    d=df.dropna(subset=[lat_col,lon_col]).copy(); d=d[d["evento_norm"].isin(EVENTOS_OPTIMIZABLES)].copy()
    if d.empty: return pd.DataFrame()
    d["_es_visita_sin_carga"]=d["evento_norm"].isin(["CSC","CNP","NHP"])
    d["id_punto_opt"]=d[col_reserva].astype(str) if col_reserva else d[lat_col].round(6).astype(str)+"_"+d[lon_col].round(6).astype(str)
    d["_os"]=pd.to_numeric(d[col_q_os],errors="coerce").fillna(1) if col_q_os else 1; d.loc[d["_os"]<=0,"_os"]=1
    d["_q_os_b"]=pd.to_numeric(d[col_q_os_b],errors="coerce").fillna(0).clip(lower=0) if col_q_os_b else 0
    d["_kilo"]=pd.to_numeric(d[col_km],errors="coerce").fillna(0).clip(lower=0) if col_km else 0
    ajuste_map=dicruta.set_index("ruta")["ajuste_vol"].to_dict() if dicruta is not None and "ajuste_vol" in dicruta.columns else {}
    d["_ajuste_ruta"]=d["ruta"].map(ajuste_map)
    aj=ajustar_volumen_m3(d[col_km],d["_os"],ajuste_vol_por_os=d["_ajuste_ruta"]) if col_km else pd.DataFrame(index=d.index)
    d["_volumen_original"]=aj.get("volumen_original_m3",0); d["_volumen"]=aj.get("volumen_m3",0); d["_m3_por_os"]=aj.get("m3_por_os",0); d["_ajuste_usado"]=aj.get("ajuste_vol_usado",0.33); d["_volumen_ajustado"]=aj.get("volumen_ajustado",False)
    # Excepciones se visitan, pero no consumen capacidad ni cupo BT.
    d.loc[d["_es_visita_sin_carga"],["_os","_q_os_b","_kilo","_volumen_original","_volumen","_m3_por_os"]]=0
    d.loc[d["_es_visita_sin_carga"],"_volumen_ajustado"]=False
    group_cols=["id_punto_opt",lat_col,lon_col,"ventana_inicio","ventana_fin"]
    opt=d.groupby(group_cols,dropna=False).agg(os=("_os","sum"),q_os_B=("_q_os_b","sum"),volumen=("_volumen","sum"),volumen_original=("_volumen_original","sum"),kilo_mayor=("_kilo","sum"),m3_por_os_max=("_m3_por_os","max"),ajuste_vol_usado=("_ajuste_usado","max"),volumen_ajustado=("_volumen_ajustado","max"),es_visita_sin_carga=("_es_visita_sin_carga","max"),datetime_gestion=("datetime_gestion","min"),ruta_original=("ruta","first"),vuelta_original=("vuelta","first")).reset_index().rename(columns={lat_col:"lat",lon_col:"lon"})
    opt["reserva"]=opt["id_punto_opt"].astype(str); opt["conteo_reserva"]=1
    opt["tw_start"]=opt["ventana_inicio"].apply(lambda x:time_to_minutes(x) if pd.notna(x) else np.nan).fillna(0).astype(int)
    opt["tw_end"]=opt["ventana_fin"].apply(lambda x:time_to_minutes(x) if pd.notna(x) else np.nan).fillna(1439).astype(int)
    mask=opt["tw_start"]>opt["tw_end"]; opt.loc[mask,["tw_start","tw_end"]]=[0,1439]
    opt["factor_escala_capacidad"]=1.0
    return opt.reset_index(drop=True)


def evaluar_compatibilidad_reserva_ruta(reserva, ruta, capacidad_disponible=None, reservas_disponibles=None, eta_min=None, fin_estimado=None):
    motivos=[]
    vol=float(reserva.get("volumen",0) or 0); bt_res=float(reserva.get("q_os_B",0) or 0)
    vmin=ruta.get("vol_min",np.nan); vmax=ruta.get("vol_max",np.nan); bt=ruta.get("bt",np.nan)
    if pd.notna(vmin) and vol < float(vmin)-1e-9: motivos.append("fuera del rango de volumen: menor a Vol-min")
    if pd.notna(vmax) and vol > float(vmax)+1e-9: motivos.append("fuera del rango de volumen: mayor a Vol-max")
    if pd.notna(bt):
        bt=float(bt)
        if bt==0 and bt_res>0: motivos.append("no cumple restricción BT: la ruta no admite Big Ticket")
        elif bt>=1 and not (bt_res==0 or bt_res>=bt): motivos.append(f"no cumple restricción BT: requiere 0 o al menos {int(bt)} Big Ticket")
    if capacidad_disponible is not None and vol>float(capacidad_disponible)+1e-9: motivos.append("supera la capacidad disponible")
    if reservas_disponibles is not None and float(reservas_disponibles)<1: motivos.append("no existe cupo de reservas")
    hmin=ruta.get("hora_min_minutos",np.nan); hmax=ruta.get("hora_max_minutos",np.nan)
    if eta_min is not None:
        if pd.notna(hmin) and eta_min < int(hmin): motivos.append("incompatibilidad horaria: inicio anterior a Hora-min")
        if pd.notna(hmax) and eta_min > int(hmax): motivos.append("incompatibilidad horaria: llegada posterior a Hora-max")
        if eta_min < int(reserva.get("tw_start",0)) or eta_min > int(reserva.get("tw_end",1439)): motivos.append("ventana horaria imposible")
    if fin_estimado is not None and pd.notna(hmax) and fin_estimado>int(hmax): motivos.append("incompatibilidad horaria: término posterior a Hora-max")
    return {"compatible":len(motivos)==0,"motivos":motivos}


def construir_flotas_baseline(baseline,dicruta):
    rutas=sorted(baseline["ruta"].dropna().unique()); cfg=dicruta.set_index("ruta").to_dict("index")
    rows=[]
    for r in rutas:
        x=cfg.get(r,{})
        rows.append({
            "vehiculo_base":r,
            "capacidad":float(x.get("capacidad",DEFAULT_CAPACIDAD) if pd.notna(x.get("capacidad",np.nan)) else DEFAULT_CAPACIDAD),
            "tipo_ruta":str(x.get("tipo_ruta","SIN INFORMACIÓN")),
            "tipo_flota":str(x.get("tipo_flota","SIN INFORMACIÓN")),
            "horario_min":x.get("hora_min_minutos",np.nan),
            "hora_min_minutos":x.get("hora_min_minutos",np.nan),
            "hora_max_minutos":x.get("hora_max_minutos",np.nan),
            "vol_min":x.get("vol_min",np.nan),
            "vol_max":x.get("vol_max",np.nan),
            "resv_max":x.get("resv_max",np.nan),
            "bt":x.get("bt",np.nan),
            "ajuste_vol":x.get("ajuste_vol",np.nan),
            "es_propio":bool(x.get("es_propio",False)),
            "config_valida":bool(x.get("config_valida",True))
        })
    return pd.DataFrame(rows)


def _construir_vehiculos_v18(flota):
    vehs=[]
    for _,r in flota.iterrows():
        hmin=int(r["hora_min_minutos"]) if pd.notna(r.get("hora_min_minutos",np.nan)) else None
        hmax=int(r["hora_max_minutos"]) if pd.notna(r.get("hora_max_minutos",np.nan)) else None
        for bloque,vuelta,bstart,bend in [("AM",1,BLOQUE_AM_START_MIN,BLOQUE_AM_END_MIN),("PM",2,BLOQUE_PM_START_MIN,BLOQUE_PM_END_MIN)]:
            s=max(bstart,hmin) if hmin is not None else bstart; e=min(bend,hmax) if hmax is not None else bend
            if s>=e or e-s<1: continue
            x=r.to_dict(); x.update({"vuelta_slot":vuelta,"bloque":bloque,"slot_start":int(s),"slot_end":int(e)})
            vehs.append(x)
    return pd.DataFrame(vehs)


def _diagnosticar_no_asignada(p, vehiculos, motivos_extra=None):
    motivos=list(motivos_extra or []); rutas_eval=[]; compatibles=0
    for _,v in vehiculos.iterrows():
        rutas_eval.append(str(v["vehiculo_base"])); ev=evaluar_compatibilidad_reserva_ruta(p,v)
        if ev["compatible"]: compatibles+=1
        else: motivos.extend(ev["motivos"])
    if not rutas_eval: motivos.append("no existe una ruta compatible")
    if rutas_eval and compatibles==0 and not motivos: motivos.append("no existe una ruta compatible")
    return {"reserva":p.get("id_punto_opt",""),"ruta_original":p.get("ruta_original",""),"volumen":p.get("volumen",0),"q_os":p.get("os",0),"q_os_B":p.get("q_os_B",0),"ventana_horaria":f"{minutes_to_time(p.get('tw_start',0))}-{minutes_to_time(p.get('tw_end',1439))}","motivos":"; ".join(sorted(set(motivos))) or "no existe una ruta compatible","rutas_evaluadas":", ".join(sorted(set(rutas_eval))),"cantidad_rutas_compatibles_encontradas":compatibles}



def _resolver_ortools_v18(puntos, vehiculos, usar_osrm_matrix=True):
    """Resuelve asignación/secuencia con capacidad, reservas, horarios y vehículos permitidos.

    Compatibilidad Python 3.13 / OR-Tools:
    evita SetAllowedVehiclesForIndex(), cuya envoltura SWIG puede fallar al
    convertir listas Python al tipo absl::Span<int const>. En su lugar se
    eliminan del dominio de VehicleVar los vehículos no compatibles.
    """
    if puntos.empty or vehiculos.empty or len(puntos) > ORTOOLS_MAX_PUNTOS:
        return None, None, None
    puntos = puntos.reset_index(drop=True).copy()
    vehiculos = vehiculos.reset_index(drop=True).copy()
    coords=[HUB_COORD]+list(zip(puntos["lat"],puntos["lon"]))
    matrix,matrix_source=get_vial_matrix(tuple(coords),usar_osrm=usar_osrm_matrix)
    manager=pywrapcp.RoutingIndexManager(len(coords),len(vehiculos),0)
    routing=pywrapcp.RoutingModel(manager)
    service=[0]+[int(math.ceil(TIEMPO_SERVICIO_BASE+float(x)*TIEMPO_SERVICIO_OS)) for x in puntos["os"]]
    def time_cb(fi,ti):
        f=manager.IndexToNode(fi); t=manager.IndexToNode(ti)
        return int(matrix[f][t]+(service[f] if f else 0))
    transit=routing.RegisterTransitCallback(time_cb); routing.SetArcCostEvaluatorOfAllVehicles(transit)

    # Costo fijo por vehículo abierto: objetivo explícito de consolidación.
    # Cada vehículo/vuelta que el solver activa paga este costo fijo además
    # del tiempo de viaje. Esto empuja al solver a preferir MENOS vehículos
    # con mayor ocupación en vez de repartir poca carga en muchas rutas,
    # que era el comportamiento anterior (solo minimizaba tiempo de viaje).
    for v in range(len(vehiculos)):
        routing.SetFixedCostOfVehicle(COSTO_FIJO_VEHICULO_ABIERTO, int(v))

    # Prioridad suave de tipos de ruta: abrir una ruta NO priorizada suma un
    # costo fijo adicional por encima del costo base de consolidación.
    # Las rutas no priorizadas siguen disponibles cuando son necesarias.
    if "es_tipo_priorizado" in vehiculos.columns and vehiculos["es_tipo_priorizado"].any():
        for v,row in vehiculos.iterrows():
            if not bool(row.get("es_tipo_priorizado",False)):
                routing.SetFixedCostOfVehicle(
                    COSTO_FIJO_VEHICULO_ABIERTO + PENALIZACION_RUTA_NO_PRIORIZADA,
                    int(v)
                )

    routing.AddDimension(transit,30,1440,False,"Time"); td=routing.GetDimensionOrDie("Time")
    for n,p in enumerate(puntos.to_dict("records"),1):
        td.CumulVar(manager.NodeToIndex(n)).SetRange(int(p["tw_start"]),int(p["tw_end"]))
    for v,row in vehiculos.iterrows():
        td.CumulVar(routing.Start(v)).SetRange(int(row["slot_start"]),int(row["slot_end"]))
        td.CumulVar(routing.End(v)).SetRange(int(row["slot_start"]),int(row["slot_end"]))
        td.SetSpanUpperBoundForVehicle(TIEMPO_MAX_RUTA_MIN,v)
    SCALE=1000
    def cap_cb(idx):
        n=manager.IndexToNode(idx); return 0 if n==0 else int(math.ceil(float(puntos.iloc[n-1]["volumen"])*SCALE))
    capidx=routing.RegisterUnaryTransitCallback(cap_cb)
    routing.AddDimensionWithVehicleCapacity(capidx,0,[int(max(1,float(c)*SCALE)) for c in vehiculos["capacidad"]],True,"Capacity")
    def count_cb(idx): return 0 if manager.IndexToNode(idx)==0 else 1
    countidx=routing.RegisterUnaryTransitCallback(count_cb)
    max_count=max(len(puntos),1)
    count_caps=[int(r["resv_max"]) if pd.notna(r.get("resv_max",np.nan)) else max_count for _,r in vehiculos.iterrows()]
    routing.AddDimensionWithVehicleCapacity(countidx,0,count_caps,True,"Reservations")
    for n,p in enumerate(puntos.to_dict("records"),1):
        allowed=[]
        for v,row in vehiculos.iterrows():
            v_int = int(v)
            if evaluar_compatibilidad_reserva_ruta(p,row)["compatible"]:
                allowed.append(v_int)

        idx = int(manager.NodeToIndex(int(n)))
        vehicle_var = routing.VehicleVar(idx)
        allowed_set = set(allowed)

        # VehicleVar también admite -1 cuando el nodo queda descartado por la
        # disyunción. Solo retiramos del dominio los vehículos incompatibles.
        for v_int in range(len(vehiculos)):
            if v_int not in allowed_set:
                vehicle_var.RemoveValue(v_int)

        # Permite descartar nodos imposibles con una penalización alta; se
        # reportan como no asignados y nunca como una ruta ficticia.
        routing.AddDisjunction([idx],10_000_000)
    params=pywrapcp.DefaultRoutingSearchParameters(); params.first_solution_strategy=routing_enums_pb2.FirstSolutionStrategy.PARALLEL_CHEAPEST_INSERTION
    params.local_search_metaheuristic=routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH; params.time_limit.seconds=max(1, ORTOOLS_TIME_LIMIT_SECONDS)
    sol=routing.SolveWithParameters(params)
    if not sol: return None,None,matrix_source
    asign=[]; asignados=set()
    for v in range(len(vehiculos)):
        idx=routing.Start(v); stops=[]
        while not routing.IsEnd(idx):
            n=manager.IndexToNode(idx)
            if n:
                p=puntos.iloc[n-1].to_dict(); asignados.add(n-1); stops.append({"p":p,"eta_min":sol.Value(td.CumulVar(idx)),"cumple_vh":True,"compatible":True,"motivos":""})
            idx=sol.Value(routing.NextVar(idx))
        if stops: asign.append({"vehiculo":vehiculos.iloc[v].to_dict(),"stops":stops,"fin_min":sol.Value(td.CumulVar(idx))})
    dropped=[puntos.iloc[i].to_dict() for i in range(len(puntos)) if i not in asignados]
    return asign,dropped,matrix_source


def optimizar_cached(puntos,flota,escenario,usar_osrm_matrix=True,usar_osrm_geometry=True):
    if puntos.empty or flota.empty:
        noas=pd.DataFrame([_diagnosticar_no_asignada(p,{ }.get('x',pd.DataFrame())) for p in puntos.to_dict('records')]) if not puntos.empty else pd.DataFrame()
        return [],pd.DataFrame(),pd.DataFrame(),{"status":"sin_puntos_o_flota","modelo":"v18","no_asignadas":noas}
    puntos=puntos.reset_index(drop=True).copy(); vehiculos=_construir_vehiculos_v18(flota)
    if vehiculos.empty:
        noas=pd.DataFrame([_diagnosticar_no_asignada(p,vehiculos) for p in puntos.to_dict('records')])
        return [],pd.DataFrame(),pd.DataFrame(),{"status":"sin_vehiculos_factibles","modelo":"v18","no_asignadas":noas}
    def travel(a,b): return int(math.ceil(haversine_km(a[0],a[1],b[0],b[1])*1.35/28*60))
    pendientes=puntos.sort_values(["tw_end","tw_start","volumen"],ascending=[True,True,False]).to_dict("records")
    asignaciones=[]
    modelo="heuristica_v18_restricciones_completas"
    ort_asig, ort_drop, matrix_source = _resolver_ortools_v18(puntos, vehiculos, usar_osrm_matrix)
    if ort_asig is not None:
        asignaciones=ort_asig; pendientes=ort_drop; modelo="ortools_v18_restricciones_completas"
    for _,v in (vehiculos.iterrows() if ort_asig is None else []):
        if not pendientes: break
        cap=float(v["capacidad"]); cupo=float(v["resv_max"]) if pd.notna(v.get("resv_max",np.nan)) else math.inf
        t=int(v["slot_start"]); pos=HUB_COORD; stops=[]
        while pendientes:
            candidatos=[]
            for i,p in enumerate(pendientes):
                viaje=travel(pos,(float(p["lat"]),float(p["lon"]))); eta=max(t+viaje,int(p["tw_start"])); serv=int(math.ceil(TIEMPO_SERVICIO_BASE+float(p["os"])*TIEMPO_SERVICIO_OS)); ret=travel((float(p["lat"]),float(p["lon"])),HUB_COORD); fin=eta+serv+ret
                ev=evaluar_compatibilidad_reserva_ruta(p,v,cap,cupo,eta,fin)
                if fin-int(v["slot_start"])>TIEMPO_MAX_RUTA_MIN: ev["motivos"].append("supera duración máxima de ruta"); ev["compatible"]=False
                if fin>int(v["slot_end"]): ev["motivos"].append("incompatibilidad horaria: excede bloque operativo"); ev["compatible"]=False
                if ev["compatible"]: candidatos.append((int(p["tw_end"]),viaje,-float(p["volumen"]),i,eta,serv))
            if not candidatos: break
            *_,idx,eta,serv=min(candidatos); p=pendientes.pop(idx); stops.append({"p":p,"eta_min":eta,"cumple_vh":True,"compatible":True,"motivos":""}); cap-=float(p["volumen"]); cupo-=1; t=eta+serv; pos=(float(p["lat"]),float(p["lon"]))
        if stops: asignaciones.append({"vehiculo":v.to_dict(),"stops":stops,"fin_min":t+travel(pos,HUB_COORD)})
    noas=pd.DataFrame([_diagnosticar_no_asignada(p,vehiculos,["no fue posible asignar con la capacidad, cupos y horarios remanentes"]) for p in pendientes])
    routes=[]; details=[]
    for rid,a in enumerate(asignaciones,1):
        v=a["vehiculo"]; coords=[HUB_COORD]; total_os=total_vol=total_k=total_bt=0
        out=[]
        for seq,it in enumerate(a["stops"],1):
            p=it["p"]; total_os+=p["os"]; total_vol+=p["volumen"]; total_k+=p["kilo_mayor"]; total_bt+=p.get("q_os_B",0); coords.append((p["lat"],p["lon"]))
            row={"escenario":escenario,"id_ruta_optimizada":rid,"vehiculo_base":v["vehiculo_base"],"vuelta":v["vuelta_slot"],"bloque":v["bloque"],"secuencia":seq,"id_punto_opt":p["id_punto_opt"],"reserva":p["id_punto_opt"],"lat":p["lat"],"lon":p["lon"],"eta":minutes_to_time(it["eta_min"]),"eta_min":it["eta_min"],"tw_start":p["tw_start"],"tw_end":p["tw_end"],"cumple_vh_estimado":True,"compatible":True,"motivos_incompatibilidad":"","os":p["os"],"q_os_B":p.get("q_os_B",0),"volumen":p["volumen"],"volumen_original":p["volumen_original"],"kilo_mayor":p["kilo_mayor"],"m3_por_os_max":p.get("m3_por_os_max",0),"ajuste_vol_usado":p.get("ajuste_vol_usado",0.33),"volumen_ajustado":p.get("volumen_ajustado",False),"ruta_real_original":p["ruta_original"],"vuelta_real_original":p["vuelta_original"]}
            out.append(row); details.append(row)
        coords.append(HUB_COORD); geom,ok,km=get_route_geometry(tuple(coords),usar_osrm=usar_osrm_geometry)
        routes.append({"escenario":escenario,"id_ruta_optimizada":rid,"vehiculo_base":v["vehiculo_base"],"vuelta":v["vuelta_slot"],"bloque":v["bloque"],"capacidad":v["capacidad"],"resv_max":v.get("resv_max",np.nan),"bt":v.get("bt",np.nan),"factor_ocupacion":total_vol/v["capacidad"]*100,"color":color_from_text(f"{escenario}-{rid}"),"geometry":geom,"geometry_ok":ok,"stops":out,"total_os":total_os,"total_q_os_B":total_bt,"volumen":total_vol,"kilo_mayor":total_k,"paradas":len(out),"reservas":len(out),"km_estimado":km,"inicio_ruta":minutes_to_time(v["slot_start"]),"fin_ruta":minutes_to_time(a["fin_min"]),"tiempo_ruta_min":a["fin_min"]-v["slot_start"],"activo":True,"factible":True})
    resumen=pd.DataFrame([{k:v for k,v in r.items() if k not in ["geometry","stops","color"]} for r in routes])
    meta={"status":"ok" if noas.empty else "ok_con_no_asignadas","matrix_source":"haversine/osrm_geometry","modelo":modelo,"no_asignadas":noas,"reservas_no_asignadas":len(noas)}
    return routes,pd.DataFrame(details),resumen,meta



def _diagnostico_baseline(baseline,dicruta,cols):
    pts=preparar_puntos(baseline,cols,dicruta=dicruta); cfg=dicruta.set_index("ruta").to_dict("index"); rows=[]
    for p in pts.to_dict("records"):
        r=cfg.get(p["ruta_original"])
        ev={"compatible":False,"motivos":["ruta original no existe en diccionario"]} if r is None else evaluar_compatibilidad_reserva_ruta(p,r)
        rows.append({"reserva":p["id_punto_opt"],"ruta":p["ruta_original"],"volumen":p["volumen"],"q_os":p["os"],"q_os_B":p.get("q_os_B",0),"cumple_config_reserva":ev["compatible"],"motivos_config_reserva":"; ".join(ev["motivos"])})
    return pd.DataFrame(rows)


def metricas_baseline(baseline,dicruta,cols):
    # Reutiliza el cálculo histórico, agregando diagnóstico sin modificar la operación real.
    d=baseline.copy(); cap_map=dicruta.set_index("ruta")["capacidad"].to_dict(); ajuste_map=dicruta.set_index("ruta")["ajuste_vol"].to_dict()
    q=pd.to_numeric(d[cols["col_q_os"]],errors="coerce").fillna(1) if cols["col_q_os"] else pd.Series(1,index=d.index)
    aj=ajustar_volumen_m3(d[cols["col_km"]],q,ajuste_vol_por_os=d["ruta"].map(ajuste_map)) if cols["col_km"] else pd.DataFrame({"volumen_m3":0},index=d.index)
    d["volumen_calc"]=aj["volumen_m3"]; d["capacidad_ruta"]=d["ruta"].map(cap_map).fillna(DEFAULT_CAPACIDAD)
    rutas=d["ruta"].nunique(); vueltas=d.dropna(subset=["vuelta"]).drop_duplicates(["ruta","vuelta"]).shape[0]; total_os=pd.to_numeric(d[cols["col_q_os"]],errors="coerce").fillna(0).sum() if cols["col_q_os"] else len(d); reservas=d[cols["col_reserva"]].nunique()
    occ=d.groupby(["ruta","vuelta"],dropna=False).agg(vol=("volumen_calc","sum"),cap=("capacidad_ruta","first")).reset_index(); occ["fo"]=np.where(occ["cap"]>0,occ["vol"]/occ["cap"]*100,np.nan)
    diag=_diagnostico_baseline(baseline,dicruta,cols); inc=int((~diag["cumple_config_reserva"]).sum()) if not diag.empty else 0
    return {"escenario":"Baseline","cantidad_rutas":rutas,"vueltas":vueltas,"total_os":total_os,"reservas":reservas,"volumen":d["volumen_calc"].sum(),"cumplimiento_vh":d["cumple_ventana"].mean()*100 if d["cumple_ventana"].notna().any() else np.nan,"excepciones":d["es_excepcion"].mean()*100 if len(d) else np.nan,"productividad":total_os/rutas if rutas else np.nan,"factor_ocupacion":occ["fo"].mean() if not occ.empty else np.nan,"ocupacion_vehiculo_dia":occ["fo"].mean() if not occ.empty else np.nan,"capacidad_total":occ["cap"].sum() if not occ.empty else np.nan,"tiempo_ruta":((d.groupby("ruta")["datetime_gestion"].max()-d.groupby("ruta")["datetime_gestion"].min()).dt.total_seconds().fillna(0).sum()/60/rutas) if rutas else np.nan,"reservas_incompatibles":inc}


# ============================================================
# SIDEBAR / INPUTS
# ============================================================
with st.sidebar:
    st.header("Carga")
    archivo_peu = st.file_uploader("Archivo gestión PU", type=["xlsx"], key="peu")
    archivo_accesos = st.file_uploader("Archivo accesos / ingreso Hub1", type=["xlsx"], key="accesos")
    archivo_dicruta = st.file_uploader("Diccionario de rutas/capacidad (opcional)", type=["xlsx"], key="dicruta")
    st.caption("Si no cargas dicruta, el aplicativo intentará leer un archivo fijo llamado dicruta.xlsx junto al .py.")
    with st.expander("Información técnica", expanded=False):
        st.caption(f"OR-Tools: hasta {ORTOOLS_MAX_PUNTOS} puntos · límite {ORTOOLS_TIME_LIMIT_SECONDS}s")
        st.caption(f"OSRM: {'deshabilitado' if DISABLE_OSRM else 'habilitado'}")
        if st.button("Limpiar caché", use_container_width=True):
            st.cache_data.clear()
            st.success("Caché limpiado.")

if archivo_peu is None or archivo_accesos is None:
    st.info("Carga los archivos de gestión PU y accesos para iniciar. El optimizador se ejecutará automáticamente con los parámetros bloqueados.")
    st.stop()

try:
    peu_raw = read_excel_any(archivo_peu)
    accesos_raw = read_excel_any(archivo_accesos)
    if archivo_dicruta is not None:
        dicruta = cargar_dicruta(archivo_dicruta)
    elif DICRUTA_DEFAULT_PATH.exists():
        dicruta = cargar_dicruta(str(DICRUTA_DEFAULT_PATH))
    else:
        st.error("No encontré dicruta.xlsx. Cárgalo en la barra lateral o déjalo fijo junto al aplicativo.")
        st.stop()
except Exception as e:
    st.error(f"Error cargando archivos: {e}")
    st.caption("Verifica que los archivos sean .xlsx válidos y que sus columnas correspondan al formato esperado.")
    st.stop()


# Preferencia opcional de tipos de ruta para la optimización de capacidades.
tipos_ruta_dic = sorted([
    str(x).strip()
    for x in dicruta.get("tipo_ruta", pd.Series(dtype=str)).dropna().unique()
    if str(x).strip() and str(x).strip().upper() != "SIN INFORMACIÓN"
])

with st.sidebar:
    st.markdown("---")
    st.subheader("Prioridad de rutas")
    tipos_ruta_priorizados = st.multiselect(
        "Tipos de ruta a priorizar",
        options=tipos_ruta_dic,
        default=[],
        help=(
            "El optimizador intentará asignar carga primero a estos tipos de ruta, "
            "siempre que cumplan capacidad, volumen, BT, cupos y horarios. "
            "Si no son suficientes, utilizará los demás tipos disponibles."
        ),
        key="tipos_ruta_priorizados"
    )
    if tipos_ruta_priorizados:
        st.caption(
            "Prioridad activa: " + ", ".join(tipos_ruta_priorizados)
        )
    else:
        st.caption("Sin prioridad: se mantiene la lógica normal por capacidad y tiempo.")


# Configuración y validaciones del diccionario v18
with st.expander("⚙️ Configuración de rutas cargada", expanded=False):
    cfg_show = dicruta[["ruta","capacidad","horario_txt","vol_min","vol_max","resv_max","bt","ajuste_vol","es_propio","config_valida"]].copy()
    cfg_show["rango_volumen"] = cfg_show.apply(lambda r: f"{r['vol_min'] if pd.notna(r['vol_min']) else '-'} a {r['vol_max'] if pd.notna(r['vol_max']) else '-'}", axis=1)
    st.dataframe(cfg_show[["ruta","capacidad","horario_txt","rango_volumen","resv_max","bt","ajuste_vol","es_propio","config_valida"]], use_container_width=True, hide_index=True)
    if DICRUTA_VALIDACIONES:
        st.warning(f"Se detectaron {len(DICRUTA_VALIDACIONES)} observaciones de configuración.")
        st.dataframe(pd.DataFrame(DICRUTA_VALIDACIONES), use_container_width=True, hide_index=True)
    else:
        st.success("Diccionario sin observaciones de validación.")

peu, accesos, cols, faltantes = preparar_data(peu_raw, accesos_raw)
if faltantes:
    st.error("Faltan columnas necesarias: " + ", ".join(faltantes)); st.stop()

clases_disponibles = sorted(peu["clase_ruta"].dropna().unique())
tipos_disponibles = sorted(peu["tipo_ruta"].dropna().unique())
# Por defecto se consideran todas las clases de ruta disponibles.
default_clases = clases_disponibles
comunas_disponibles = sorted(peu["comuna"].dropna().unique())
default_comuna = "Todas"



def obtener_rutas_propias_detectadas(df_operacion, dicruta):
    """Devuelve las rutas propias presentes en la operación filtrada."""
    if df_operacion is None or df_operacion.empty or dicruta is None or dicruta.empty:
        return []
    if "es_propio" not in dicruta.columns:
        return []
    propias_dic = set(dicruta.loc[dicruta["es_propio"], "ruta"].dropna().astype(str))
    rutas_operacion = set(df_operacion["ruta"].dropna().astype(str))
    return sorted(propias_dic.intersection(rutas_operacion))


def construir_configuracion_propios(df_operacion, dicruta, key_prefix="propios_global"):
    """Muestra las decisiones previas para reservas y vehículos propios."""
    rutas_detectadas = obtener_rutas_propias_detectadas(df_operacion, dicruta)

    config_default = {
        "rutas_propias_detectadas": rutas_detectadas,
        "optimizar_reservas_propias": False,
        "vehiculos_propios_incluidos": []
    }

    if not rutas_detectadas:
        st.success("No se detectaron rutas propias en la operación filtrada.")
        return config_default

    detalle = (
        dicruta[dicruta["ruta"].isin(rutas_detectadas)]
        [["ruta", "capacidad", "tipo_flota"]]
        .drop_duplicates("ruta")
        .sort_values("ruta")
        .reset_index(drop=True)
    )

    st.warning(
        "Se detectaron rutas propias en la operación filtrada. "
        "Define cómo deben tratarse antes de ejecutar los escenarios optimizados."
    )
    st.dataframe(
        detalle.rename(columns={
            "ruta": "Ruta / vehículo propio",
            "capacidad": "Capacidad m³",
            "tipo_flota": "Tipo de flota"
        }),
        use_container_width=True,
        hide_index=True
    )

    optimizar_reservas = st.radio(
        "¿Deseas incluir las reservas asociadas a estas rutas propias en la optimización?",
        ["No", "Sí"],
        horizontal=True,
        index=0,
        key=f"{key_prefix}_optimizar_reservas"
    ) == "Sí"

    vehiculos_incluidos = []
    if optimizar_reservas:
        incluir_flota = st.radio(
            "¿Deseas incorporar vehículos propios dentro de las opciones de flota de la optimización?",
            ["No", "Sí"],
            horizontal=True,
            index=0,
            key=f"{key_prefix}_incluir_flota"
        ) == "Sí"

        if incluir_flota:
            detalle["opcion"] = detalle.apply(
                lambda r: f"{r['ruta']} · {float(r['capacidad']):.1f} m³ · {r['tipo_flota']}",
                axis=1
            )
            mapa = dict(zip(detalle["opcion"], detalle["ruta"]))
            seleccion = st.multiselect(
                "Selecciona los vehículos propios que participarán en la optimización",
                options=detalle["opcion"].tolist(),
                default=detalle["opcion"].tolist(),
                key=f"{key_prefix}_vehiculos_seleccionados"
            )
            vehiculos_incluidos = [mapa[x] for x in seleccion]

    return {
        "rutas_propias_detectadas": rutas_detectadas,
        "optimizar_reservas_propias": optimizar_reservas,
        "vehiculos_propios_incluidos": vehiculos_incluidos
    }

def comparar_cantidad_por_tipo_ruta(baseline, resumen_cap, dicruta):
    """Compara rutas físicas únicas por tipo entre Baseline y Capacidades."""
    tipo_map = (
        dicruta.set_index("ruta")["tipo_ruta"].to_dict()
        if "tipo_ruta" in dicruta.columns
        else {}
    )

    # Baseline: una ruta física se cuenta una sola vez, aunque tenga varias vueltas.
    base_rutas = baseline[["ruta"]].dropna().drop_duplicates().copy()
    base_rutas["tipo_ruta"] = (
        base_rutas["ruta"]
        .map(tipo_map)
        .fillna("SIN INFORMACIÓN")
        .astype(str)
        .str.strip()
    )
    base_count = (
        base_rutas.groupby("tipo_ruta")["ruta"]
        .nunique()
        .rename("Baseline")
    )

    # Optimización: solo vehículos/rutas efectivamente activados.
    if resumen_cap is None or resumen_cap.empty:
        opt_count = pd.Series(dtype=float, name="Optimización de capacidades")
    else:
        opt = resumen_cap.copy()
        if "activo" in opt.columns:
            opt = opt[opt["activo"].fillna(True).astype(bool)].copy()
        elif "paradas" in opt.columns:
            opt = opt[
                pd.to_numeric(opt["paradas"], errors="coerce").fillna(0) > 0
            ].copy()

        opt = opt[["vehiculo_base"]].dropna().drop_duplicates().copy()
        opt["tipo_ruta"] = (
            opt["vehiculo_base"]
            .map(tipo_map)
            .fillna("SIN INFORMACIÓN")
            .astype(str)
            .str.strip()
        )
        opt_count = (
            opt.groupby("tipo_ruta")["vehiculo_base"]
            .nunique()
            .rename("Optimización de capacidades")
        )

    comparacion = pd.concat([base_count, opt_count], axis=1).fillna(0).reset_index()
    comparacion["Baseline"] = comparacion["Baseline"].astype(int)
    comparacion["Optimización de capacidades"] = (
        comparacion["Optimización de capacidades"].astype(int)
    )
    comparacion["Variación"] = (
        comparacion["Optimización de capacidades"] - comparacion["Baseline"]
    )
    comparacion = comparacion.rename(columns={"tipo_ruta": "Tipo de ruta"})
    return comparacion.sort_values(
        ["Baseline", "Optimización de capacidades", "Tipo de ruta"],
        ascending=[False, False, True]
    ).reset_index(drop=True)


# ============================================================
# EXPORTACIÓN PROFESIONAL A EXCEL
# ============================================================
def construir_resumen_ejecutivo_export(sims):
    """Arma una fila ejecutiva por día exportado: rutas, ocupación,
    productividad y reservas fuera, baseline vs optimización de
    capacidades, más el motivo de bloqueo más frecuente del día.
    """
    filas = []
    for fecha, sim_tmp in sims:
        macro = sim_tmp.get("metrics", pd.DataFrame())
        if macro is None or macro.empty:
            continue
        base = macro[macro["escenario"].eq("Baseline")]
        cap = macro[macro["escenario"].eq("Optimización de capacidades")]
        if base.empty or cap.empty:
            continue
        base, cap = base.iloc[0], cap.iloc[0]

        no_asig = sim_tmp.get("no_asignadas_cap", pd.DataFrame())
        motivo_top = "-"
        if isinstance(no_asig, pd.DataFrame) and not no_asig.empty and "motivos" in no_asig.columns:
            serie_motivos = no_asig["motivos"].astype(str).str.split("; ").explode()
            if not serie_motivos.empty:
                motivo_top = serie_motivos.value_counts().idxmax()

        filas.append({
            "fecha": fecha,
            "rutas_baseline": base.get("cantidad_rutas"),
            "rutas_optimizado": cap.get("cantidad_rutas"),
            "rutas_menos": (base.get("cantidad_rutas", 0) or 0) - (cap.get("cantidad_rutas", 0) or 0),
            "ocupacion_baseline_pct": base.get("factor_ocupacion"),
            "ocupacion_optimizado_pct": cap.get("factor_ocupacion"),
            "ocupacion_delta_pts": (cap.get("factor_ocupacion", 0) or 0) - (base.get("factor_ocupacion", 0) or 0),
            "productividad_baseline": base.get("productividad"),
            "productividad_optimizado": cap.get("productividad"),
            "volumen_m3": cap.get("volumen"),
            "reservas_no_asignadas": len(no_asig) if isinstance(no_asig, pd.DataFrame) else 0,
            "motivo_bloqueo_mas_frecuente": motivo_top,
        })
    return pd.DataFrame(filas)


def _autofit_columnas(ws, max_width=60):
    for columna in ws.columns:
        celdas_validas = [c for c in columna if c.value is not None]
        if not celdas_validas:
            continue
        largo = max(len(str(c.value)) for c in celdas_validas)
        letra = celdas_validas[0].column_letter
        ws.column_dimensions[letra].width = min(max(largo + 2, 10), max_width)


def formatear_libro_excel(workbook):
    """Aplica formato corporativo básico a todas las hojas: encabezado en
    negrita con fondo, panel congelado en la fila de encabezado y ancho de
    columna ajustado al contenido. No cambia ningún dato, solo presentación.
    El orden de las hojas ya queda correcto por el orden en que se escriben
    (resumen_ejecutivo primero, glosario al final); esta función no reordena.
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    encabezado_font = Font(bold=True, color="FFFFFF")
    encabezado_fill = PatternFill("solid", fgColor="1D2939")

    for ws in workbook.worksheets:
        if ws.max_row == 0 or ws.max_column == 0:
            continue
        for celda in ws[1]:
            celda.font = encabezado_font
            celda.fill = encabezado_fill
            celda.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        _autofit_columnas(ws)


GLOSARIO_EXPORT = [
    {"Hoja": "resumen_ejecutivo", "Descripción": "Una fila por día exportado: rutas, ocupación y productividad baseline vs. optimización de capacidades, y el motivo de bloqueo más frecuente."},
    {"Hoja": "macro_escenarios", "Descripción": "Métricas macro completas por día y escenario (Baseline / Optimización de capacidades)."},
    {"Hoja": "comparacion_tipos_ruta", "Descripción": "Cantidad de rutas físicas únicas por tipo de ruta, Baseline vs. Optimización de capacidades."},
    {"Hoja": "baseline_detalle", "Descripción": "Detalle de gestiones (una fila por evento) de la operación real del día."},
    {"Hoja": "baseline_resumen_ruta", "Descripción": "Resumen operacional agregado por ruta física en el Baseline."},
    {"Hoja": "baseline_resumen_vuelta", "Descripción": "Resumen operacional agregado por ruta y vuelta (AM/PM) en el Baseline."},
    {"Hoja": "baseline_diagnostico", "Descripción": "Hallazgos de calidad de datos detectados en el Baseline (ventanas horarias, geocodificación, etc.)."},
    {"Hoja": "opt_cap_resumen", "Descripción": "Resumen por vehículo/vuelta generado por el optimizador de capacidades."},
    {"Hoja": "opt_cap_detalle", "Descripción": "Detalle parada por parada de cada ruta generada por el optimizador de capacidades."},
    {"Hoja": "demanda_optimizable", "Descripción": "Puntos de demanda (reservas agrupadas) que se entregaron al optimizador."},
    {"Hoja": "reservas_no_asignadas", "Descripción": "Reservas que el optimizador no pudo asignar a ninguna ruta, con el o los motivos (restricción incumplida)."},
    {"Hoja": "validaciones_dicruta", "Descripción": "Filas del diccionario de rutas con configuración inválida o inconsistente detectada al cargar el archivo."},
    {"Hoja": "dicruta_configuracion", "Descripción": "Copia del diccionario de rutas/capacidad utilizado en esta corrida, tal como fue interpretado por la app."},
]


# ============================================================
# EJECUCIÓN AUTOMÁTICA DE ESCENARIOS
# ============================================================
def ejecutar_escenarios_dia(
    fecha,
    clases_sel,
    tipos_sel,
    comuna_sel=None,
    optimizar_reservas_propias=False,
    vehiculos_propios_incluidos=None,
    tipos_ruta_priorizados=None
):
    """Ejecuta baseline y optimización de capacidades.

    Reglas para rutas propias:
    - El baseline siempre conserva toda la operación real.
    - Si optimizar_reservas_propias=False, sus reservas se excluyen de la optimización.
    - En capacidades, sólo los vehículos propios seleccionados quedan disponibles como flota.
    """
    vehiculos_propios_incluidos = vehiculos_propios_incluidos or []
    tipos_ruta_priorizados = tipos_ruta_priorizados or []

    base = peu[
        (peu["fecha_operacion"] == fecha) &
        peu["clase_ruta"].isin(clases_sel) &
        peu["tipo_ruta"].isin(tipos_sel)
    ].copy()

    if comuna_sel and comuna_sel != "Todas":
        base = base[base["comuna"].eq(comuna_sel)].copy()

    acc = accesos[
        accesos["solo_fch"] == fecha
    ].sort_values(["ruta", "datetime_ingreso"]).copy()

    baseline_completo = asignar_vueltas_por_cierre(base, acc)

    propio_map = (
        dicruta.set_index("ruta")["es_propio"].to_dict()
        if "es_propio" in dicruta.columns
        else {}
    )

    baseline_completo["es_ruta_propia"] = (
        baseline_completo["ruta"]
        .map(propio_map)
        .fillna(False)
        .astype(bool)
    )

    rutas_propias_dia = sorted(
        baseline_completo.loc[
            baseline_completo["es_ruta_propia"],
            "ruta"
        ].dropna().unique()
    )

    # Demanda autorizada para optimización de capacidades.
    if optimizar_reservas_propias:
        baseline_cap = baseline_completo.copy()
    else:
        baseline_cap = baseline_completo[
            ~baseline_completo["es_ruta_propia"]
        ].copy()

    # Flota real completa del día.
    flota_bl_completa = construir_flotas_baseline(
        baseline_completo,
        dicruta
    )

    flota_bl_completa["es_propio"] = (
        flota_bl_completa["vehiculo_base"]
        .map(propio_map)
        .fillna(False)
        .astype(bool)
    )

    # Flota de capacidades: terceros + propios seleccionados.
    flota_cap_base = flota_bl_completa[
        ~flota_bl_completa["es_propio"]
    ].copy()

    if vehiculos_propios_incluidos:
        flota_prop_sel = flota_bl_completa[
            flota_bl_completa["vehiculo_base"].isin(
                vehiculos_propios_incluidos
            )
        ].copy()

        flota_cap_base = (
            pd.concat(
                [flota_cap_base, flota_prop_sel],
                ignore_index=True
            )
            .drop_duplicates("vehiculo_base")
        )

    max_capacidad_disponible = (
        float(flota_cap_base["capacidad"].max())
        if not flota_cap_base.empty
        else DEFAULT_CAPACIDAD
    )

    puntos_cap = preparar_puntos(
        baseline_cap,
        cols,
        max_capacidad=max_capacidad_disponible,
        dicruta=dicruta
    )

    flota_cap = construir_flotas_capacidades(
        flota_cap_base,
        puntos_cap["volumen"].sum()
        if not puntos_cap.empty
        else 0,
        tipos_ruta_priorizados=tipos_ruta_priorizados
    )

    routes_cap, detail_cap, resumen_cap, meta_cap = optimizar_cached(
        puntos_cap,
        flota_cap,
        "Optimización de capacidades",
        True,
        True
    )

    no_asignadas_cap = meta_cap.get(
        "no_asignadas",
        pd.DataFrame()
    )

    diagnostico_baseline = _diagnostico_baseline(
        baseline_completo,
        dicruta,
        cols
    )

    comparacion_tipos_ruta = comparar_cantidad_por_tipo_ruta(
        baseline_completo,
        resumen_cap,
        dicruta
    )

    met_base = metricas_baseline(
        baseline_completo,
        dicruta,
        cols
    )

    mets = pd.DataFrame([
        met_base,
        metricas_opt(
            resumen_cap,
            puntos_cap,
            "Optimización de capacidades"
        )
    ])

    mets.loc[
        mets["escenario"].eq("Optimización de capacidades"),
        "cumplimiento_vh"
    ] = 100.0

    mets.insert(0, "fecha", fecha)
    mets.insert(
        1,
        "comuna",
        comuna_sel if comuna_sel else "Todas"
    )

    mets["reservas_propias_optimizadas"] = bool(
        optimizar_reservas_propias
    )

    mets["vehiculos_propios_habilitados"] = (
        ", ".join(vehiculos_propios_incluidos)
        if vehiculos_propios_incluidos
        else "Ninguno"
    )
    mets["tipos_ruta_priorizados"] = (
        ", ".join(tipos_ruta_priorizados)
        if tipos_ruta_priorizados
        else "Sin prioridad"
    )

    escala_usada = (
        float(
            puntos_cap["factor_escala_capacidad"].iloc[0]
        )
        if (
            not puntos_cap.empty and
            "factor_escala_capacidad" in puntos_cap.columns
        )
        else 1.0
    )

    return dict(
        baseline=baseline_completo,
        puntos=puntos_cap,
        metrics=mets,
        routes_cap=routes_cap,
        detail_cap=detail_cap,
        resumen_cap_opt=resumen_cap,
        meta_cap=meta_cap,
        escala_capacidad=escala_usada,
        rutas_propias_detectadas=rutas_propias_dia,
        optimizar_reservas_propias=bool(
            optimizar_reservas_propias
        ),
        vehiculos_propios_incluidos=list(
            vehiculos_propios_incluidos
        ),
        tipos_ruta_priorizados=list(
            tipos_ruta_priorizados
        ),
        no_asignadas_cap=no_asignadas_cap,
        diagnostico_baseline=diagnostico_baseline,
        comparacion_tipos_ruta=comparacion_tipos_ruta,
        validaciones_dicruta=pd.DataFrame(
            DICRUTA_VALIDACIONES
        )
    )

# ============================================================
# FILTROS GENERALES
# ============================================================
st.markdown('<div class="section-title">Filtros generales de ejecución</div>', unsafe_allow_html=True)
st.caption("Estos filtros aplican a Análisis diario, Datos macro, Análisis por ruta y Exportar.")
fg1, fg2, fg3 = st.columns(3)
with fg1:
    filtro_clases = excel_filter("Clase ruta", clases_disponibles, "global_clase", default_clases)
with fg2:
    filtro_tipos = excel_filter("Tipo ruta", tipos_disponibles, "global_tipo", tipos_disponibles)
with fg3:
    filtro_comuna = excel_single_filter("Comuna", ["Todas"] + comunas_disponibles, "global_comuna", default_comuna)


# ============================================================
# CONFIGURACIÓN DE RUTAS PROPIAS
# ============================================================
operacion_filtrada_propios = peu[
    peu["clase_ruta"].isin(filtro_clases) &
    peu["tipo_ruta"].isin(filtro_tipos)
].copy()
if filtro_comuna and filtro_comuna != "Todas":
    operacion_filtrada_propios = operacion_filtrada_propios[
        operacion_filtrada_propios["comuna"].eq(filtro_comuna)
    ].copy()

rutas_propias_visibles = obtener_rutas_propias_detectadas(
    operacion_filtrada_propios,
    dicruta
)

if rutas_propias_visibles:
    titulo_propios = f"🚚 Tratamiento de rutas propias · {len(rutas_propias_visibles)} detectada(s)"
else:
    titulo_propios = "🚚 Tratamiento de rutas propias · Sin rutas detectadas"

with st.expander(titulo_propios, expanded=False):
    st.caption(
        "Define si las reservas y los vehículos propios participarán en la optimización de capacidades. "
        "Esta decisión se aplica a todos los días ejecutados con los filtros actuales."
    )
    config_propios = construir_configuracion_propios(
        operacion_filtrada_propios,
        dicruta,
        key_prefix="config_propios_global"
    )

optimizar_reservas_propias = config_propios["optimizar_reservas_propias"]
vehiculos_propios_incluidos = config_propios["vehiculos_propios_incluidos"]

# ============================================================
# TABS
# ============================================================
tab_analisis, tab_macro, tab_rutas, tab_exportar = st.tabs(["📈 Análisis diario", "📌 Datos macro diarios", "🗺️ Análisis por ruta", "⬇️ Exportar"])

with tab_analisis:
    st.markdown('<div class="section-title">Análisis diario comparativo</div>', unsafe_allow_html=True)
    st.caption("Selecciona una variable para revisar la evolución diaria entre escenarios. El contexto operacional se mantiene como referencia del día.")
    clases_sel = filtro_clases
    tipos_sel = filtro_tipos
    comuna_sel = filtro_comuna
    st.caption(f"Filtro aplicado · Comuna: {comuna_sel}")
    indicadores = ["cumplimiento_vh", "cantidad_rutas", "factor_ocupacion", "productividad", "tiempo_ruta"]
    nombres = {"cumplimiento_vh":"% Cumplimiento VH", "cantidad_rutas":"Cantidad de rutas", "factor_ocupacion":"% Ocupación", "productividad":"Productividad", "tiempo_ruta":"Tiempo en ruta"}
    indicador_sel = excel_single_filter("Variable", [nombres[i] for i in indicadores], "ana_var", nombres["productividad"])
    inv = {v:k for k,v in nombres.items()}; var = inv[indicador_sel]

    peu_fechas = peu[(peu["clase_ruta"].isin(clases_sel)) & (peu["tipo_ruta"].isin(tipos_sel))].copy()
    if comuna_sel and comuna_sel != "Todas":
        peu_fechas = peu_fechas[peu_fechas["comuna"].eq(comuna_sel)]
    fechas = sorted(peu_fechas["fecha_operacion"].dropna().unique())
    rows=[]
    prog = st.progress(0, text="Ejecutando escenarios diarios automáticamente...") if fechas else None
    for i, f in enumerate(fechas):
        res = ejecutar_escenarios_dia(
            f, clases_sel, tipos_sel, comuna_sel,
            optimizar_reservas_propias=optimizar_reservas_propias,
            vehiculos_propios_incluidos=vehiculos_propios_incluidos,
            tipos_ruta_priorizados=tipos_ruta_priorizados
        )
        rows.append(res["metrics"])
        if prog: prog.progress((i+1)/len(fechas), text=f"Escenarios calculados: {i+1}/{len(fechas)}")
    if prog: prog.empty()
    diario = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    st.session_state["diario_comparativo"] = diario

    if diario.empty:
        st.warning("No hay datos para graficar.")
    else:
        chart_df = diario[["fecha","escenario",var]].rename(columns={var:"valor"})
        chart_df["fecha"] = pd.to_datetime(chart_df["fecha"])
        chart = alt.Chart(chart_df).mark_line(point=True, strokeWidth=3).encode(
            x=alt.X("fecha:T", title="Día operacional"), y=alt.Y("valor:Q", title=indicador_sel),
            color=alt.Color("escenario:N", title="Escenario"), tooltip=["fecha:T", "escenario:N", alt.Tooltip("valor:Q", format=".2f")]
        ).properties(height=320).interactive()
        st.altair_chart(chart, use_container_width=True)

        st.markdown("#### Contexto operacional del día")
        cols_contexto = ["fecha", "total_os", "reservas", "volumen"]
        if "capacidad_total" in diario.columns:
            cols_contexto.append("capacidad_total")
        contexto = diario[diario["escenario"].eq("Baseline")][cols_contexto].copy()
        contexto = contexto.melt(
            id_vars=["fecha"],
            value_vars=[c for c in cols_contexto if c != "fecha"],
            var_name="indicador",
            value_name="valor"
        )
        contexto["fecha"] = pd.to_datetime(contexto["fecha"])
        contexto["indicador"] = contexto["indicador"].map({
            "total_os": "Total OS",
            "reservas": "Reservas",
            "volumen": "Volumen operacional (m³)",
            "capacidad_total": "Capacidad utilizada/referencial (m³)"
        })

        charts_contexto = []
        orden_indicadores = ["Reservas", "Total OS", "Volumen operacional (m³)", "Capacidad utilizada/referencial (m³)"]
        for indicador in orden_indicadores:
            df_ind = contexto[contexto["indicador"].eq(indicador)].copy()
            if df_ind.empty:
                continue
            charts_contexto.append(
                alt.Chart(df_ind).mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4).encode(
                    x=alt.X("fecha:T", title="Día operacional"),
                    y=alt.Y("valor:Q", title=indicador),
                    tooltip=["fecha:T", "indicador:N", alt.Tooltip("valor:Q", format=".2f")]
                ).properties(height=165, title=indicador)
            )
        if charts_contexto:
            st.altair_chart(alt.vconcat(*charts_contexto).resolve_scale(y="independent"), use_container_width=True)

with tab_macro:
    st.markdown('<div class="section-title">Datos macro diarios</div>', unsafe_allow_html=True)
    clases_dia = filtro_clases
    tipos_dia = filtro_tipos
    comuna_dia = filtro_comuna
    st.caption(f"Filtro aplicado · Comuna: {comuna_dia}")
    peu_fechas_dia = peu[(peu["clase_ruta"].isin(clases_dia)) & (peu["tipo_ruta"].isin(tipos_dia))].copy()
    if comuna_dia and comuna_dia != "Todas":
        peu_fechas_dia = peu_fechas_dia[peu_fechas_dia["comuna"].eq(comuna_dia)]
    fechas_dia = sorted(peu_fechas_dia["fecha_operacion"].dropna().unique())
    fecha_sel = excel_single_filter("Día", fechas_dia, "fecha_macro", fechas_dia[0] if fechas_dia else None)
    if not fecha_sel:
        st.warning("No hay fechas disponibles."); st.stop()
    with st.spinner("Calculando baseline y simulaciones automáticamente..."):
        sim = ejecutar_escenarios_dia(
            fecha_sel, clases_dia, tipos_dia, comuna_dia,
            optimizar_reservas_propias=optimizar_reservas_propias,
            vehiculos_propios_incluidos=vehiculos_propios_incluidos,
            tipos_ruta_priorizados=tipos_ruta_priorizados
        )
    st.session_state["sim_dia"] = sim

    fecha_mostrar = pd.to_datetime(fecha_sel).strftime("%d-%m-%Y")
    met_dia = sim["metrics"][sim["metrics"]["escenario"].eq("Baseline")].iloc[0]
    st.markdown(f'''
    <div class="day-card">
      <div class="day-card-title">Día operacional seleccionado</div>
      <div class="day-card-date">{fecha_mostrar}</div>
      <div class="day-card-detail">
        Comuna: <b>{comuna_dia}</b> · Reservas: <b>{num(met_dia.get('reservas'))}</b> · OS: <b>{num(met_dia.get('total_os'))}</b> · Volumen: <b>{dec(met_dia.get('volumen'))} m³</b> · Rutas baseline: <b>{num(met_dia.get('cantidad_rutas'))}</b>
      </div>
    </div>
    ''', unsafe_allow_html=True)
    st.info("Baseline representa la operación real. La optimización de capacidades permite comparar mejoras sobre la misma demanda operativa disponible.")
    if tipos_ruta_priorizados:
        st.caption(
            "Tipos de ruta priorizados en esta simulación: "
            + ", ".join(tipos_ruta_priorizados)
        )
    estados_ok = {"ok", "ok_con_no_asignadas"}
    if sim.get("meta_cap", {}).get("status") not in estados_ok:
        st.warning("No fue posible generar la optimización de capacidades con los filtros seleccionados. Revisa la demanda disponible y vuelve a intentar.")
    else:
        st.success("Escenarios generados correctamente.")

    # ==========================================================
    # RESUMEN EJECUTIVO
    # Responde en una sola vista las preguntas que la operación
    # necesita resolver rápido: qué mejoró, cuánto, y qué lo frenó.
    # ==========================================================
    st.markdown('<div class="section-title">Resumen ejecutivo</div>', unsafe_allow_html=True)

    _macro_exec = sim["metrics"].copy()
    _base_exec = _macro_exec[_macro_exec["escenario"].eq("Baseline")].iloc[0]
    _cap_exec = _macro_exec[_macro_exec["escenario"].eq("Optimización de capacidades")].iloc[0]

    _delta_rutas = float(_cap_exec["cantidad_rutas"]) - float(_base_exec["cantidad_rutas"])
    _delta_ocupacion = float(_cap_exec["factor_ocupacion"]) - float(_base_exec["factor_ocupacion"])
    _no_asignadas_df = sim.get("no_asignadas_cap", pd.DataFrame())
    _n_no_asignadas = len(_no_asignadas_df)

    if _delta_rutas < 0 and _delta_ocupacion >= 0:
        _veredicto = f"✅ La optimización de capacidades usa **{abs(_delta_rutas):.0f} ruta(s) menos** y sube la ocupación **{_delta_ocupacion:+.1f} pts**, sobre el mismo volumen operado."
    elif _delta_rutas < 0:
        _veredicto = f"✅ La optimización de capacidades usa **{abs(_delta_rutas):.0f} ruta(s) menos**, aunque la ocupación promedio varió {_delta_ocupacion:+.1f} pts."
    elif _delta_rutas == 0 and _delta_ocupacion > 0:
        _veredicto = f"✅ Con la misma cantidad de rutas, la ocupación promedio sube **{_delta_ocupacion:+.1f} pts** (menos capacidad ociosa)."
    else:
        _veredicto = "ℹ️ Con la demanda y restricciones de este día, el baseline ya está cerca del óptimo alcanzable; revisa las restricciones que limitaron más mejora abajo."

    st.markdown(f"""
<div class="day-card" style="background:#F0FDF4;border-color:#BBF7D0;">
  <div class="day-card-title">¿Qué mejoró hoy?</div>
  <div style="font-size:1.05rem;color:#101828;margin-top:.3rem;">{_veredicto}</div>
</div>
""", unsafe_allow_html=True)

    _delta_productividad = float(_cap_exec["productividad"]) - float(_base_exec["productividad"])

    e1, e2, e3, e4 = st.columns(4)
    e1.metric(
        "Rutas: baseline → optimizado",
        f"{num(_base_exec['cantidad_rutas'])} → {num(_cap_exec['cantidad_rutas'])}",
        delta=f"{_delta_rutas:+.0f} rutas",
        delta_color="inverse"
    )
    e2.metric(
        "Ocupación: baseline → optimizado",
        f"{pct(_base_exec['factor_ocupacion'])} → {pct(_cap_exec['factor_ocupacion'])}",
        delta=f"{_delta_ocupacion:+.1f} pts"
    )
    e3.metric(
        "Productividad: baseline → optimizado",
        f"{dec(_base_exec['productividad'])} → {dec(_cap_exec['productividad'])}",
        delta=f"{_delta_productividad:+.1f}"
    )
    e4.metric(
        "Reservas que quedaron fuera",
        num(_n_no_asignadas),
        delta=None if _n_no_asignadas == 0 else f"{_n_no_asignadas} sin asignar",
        delta_color="inverse"
    )

    with st.expander("¿Qué restricciones impidieron mejorar aún más?", expanded=_n_no_asignadas > 0):
        if _no_asignadas_df.empty:
            st.success("No hubo reservas bloqueadas por restricciones: toda la demanda optimizable fue asignada.")
        else:
            _motivos_exp = (
                _no_asignadas_df["motivos"]
                .astype(str)
                .str.split("; ")
                .explode()
                .value_counts()
                .reset_index()
            )
            _motivos_exp.columns = ["Restricción / motivo", "Reservas afectadas"]
            st.caption(
                f"{_n_no_asignadas} reserva(s) quedaron fuera de la optimización de capacidades. "
                "Estos son los motivos más frecuentes (una reserva puede tener más de un motivo):"
            )
            st.dataframe(_motivos_exp, use_container_width=True, hide_index=True)

    d1, d2 = st.columns(2)
    d1.metric(
        "Reservas no asignadas",
        len(sim.get("no_asignadas_cap", pd.DataFrame()))
    )
    d2.metric(
        "Rutas con configuración inválida",
        int((~dicruta["config_valida"]).sum())
    )

    with st.expander("Reservas no asignadas", expanded=False):
        st.markdown(
            "##### Optimización de capacidades · Reservas no asignadas"
        )
        st.dataframe(
            sim.get("no_asignadas_cap", pd.DataFrame()),
            use_container_width=True,
            hide_index=True
        )

    macro = sim["metrics"].copy()
    cols_show = ["escenario","cantidad_rutas","total_os","cumplimiento_vh","vueltas","volumen","capacidad_total","excepciones","reservas","productividad","factor_ocupacion","ocupacion_vehiculo_dia","tiempo_ruta"]
    pct_cols = ["cumplimiento_vh", "excepciones", "factor_ocupacion", "ocupacion_vehiculo_dia"]
    fmt = {c: "{:.1f}%" for c in pct_cols if c in macro.columns}
    fmt.update({"productividad": "{:.1f}", "tiempo_ruta": "{:.1f}", "volumen": "{:.3f}", "capacidad_total": "{:.3f}", "total_os":"{:.0f}", "reservas":"{:.0f}"})

    st.markdown("#### Cuadro comparativo")
    card_cols = st.columns(2)
    for idx, (_, row) in enumerate(macro.iterrows()):
        with card_cols[idx]:
            st.markdown(f"""
<div style="border:1px solid #EAECF0;border-radius:16px;padding:14px;background:#FFFFFF;min-height:270px;">
  <div style="font-size:1.05rem;font-weight:800;margin-bottom:8px;">{row['escenario']}</div>
  <div style="font-size:0.85rem;color:#667085;margin-bottom:12px;">Resultados del día seleccionado</div>
  <div><b>Rutas:</b> {num(row['cantidad_rutas'])}</div>
  <div><b>Vueltas:</b> {num(row['vueltas'])}</div>
  <div><b>Cumplimiento VH:</b> {pct(row['cumplimiento_vh'])}</div>
  <div><b>Ocupación:</b> {pct(row['factor_ocupacion'])}</div>
  <div><b>Productividad:</b> {dec(row['productividad'])}</div>
  <div><b>Tiempo ruta:</b> {dec(row['tiempo_ruta'])} min</div>
  <hr style="border:none;border-top:1px solid #EAECF0;margin:10px 0;">
  <div><b>OS:</b> {num(row['total_os'])}</div>
  <div><b>Reservas:</b> {num(row['reservas'])}</div>
  <div><b>Volumen m³:</b> {dec(row['volumen'])}</div>
</div>
""", unsafe_allow_html=True)

    st.markdown("#### Tabla detalle")
    st.dataframe(macro[cols_show].style.format(fmt, na_rep="-"), use_container_width=True, hide_index=True)

    st.caption("Baseline muestra el contexto completo. La optimización de capacidades muestra demanda PU optimizable; por eso reservas/OS pueden bajar si había excepciones o puntos sin coordenadas.")

    st.markdown("#### Comparación de cantidad por tipo de ruta")
    st.caption(
        "Se cuentan rutas físicas únicas. Una ruta con vuelta AM y PM se contabiliza una sola vez."
    )

    comp_tipos = sim.get("comparacion_tipos_ruta", pd.DataFrame()).copy()

    if comp_tipos.empty:
        st.info("No hay información de tipos de ruta para comparar.")
    else:
        st.dataframe(
            comp_tipos,
            use_container_width=True,
            hide_index=True
        )

        comp_chart = comp_tipos.melt(
            id_vars=["Tipo de ruta"],
            value_vars=["Baseline", "Optimización de capacidades"],
            var_name="Escenario",
            value_name="Cantidad de rutas"
        )

        grafico_tipos = (
            alt.Chart(comp_chart)
            .mark_bar()
            .encode(
                x=alt.X(
                    "Tipo de ruta:N",
                    title="Tipo de ruta",
                    sort="-y"
                ),
                y=alt.Y(
                    "Cantidad de rutas:Q",
                    title="Cantidad de rutas",
                    axis=alt.Axis(tickMinStep=1)
                ),
                xOffset="Escenario:N",
                color=alt.Color(
                    "Escenario:N",
                    title="Escenario"
                ),
                tooltip=[
                    "Tipo de ruta:N",
                    "Escenario:N",
                    alt.Tooltip(
                        "Cantidad de rutas:Q",
                        format=".0f"
                    )
                ]
            )
            .properties(height=320)
        )

        st.altair_chart(
            grafico_tipos,
            use_container_width=True
        )

with tab_rutas:
    st.markdown(
        '<div class="section-title">Análisis por ruta / mapas OSRM</div>',
        unsafe_allow_html=True
    )
    st.caption(f"Filtro aplicado · Comuna: {filtro_comuna}")

    sim = st.session_state.get("sim_dia")
    if sim is None:
        st.warning(
            "Primero entra a Datos macro diarios para seleccionar el día."
        )
        st.stop()

    sub1, sub2, sub3, sub4 = st.tabs([
        "Baseline",
        "Optimización de capacidades",
        "Tablas",
        "Ranking de ocupación"
    ])

    with sub1:
        stops_bl, paths_bl = preparar_mapa_baseline(
            sim["baseline"],
            cols
        )

        if not stops_bl.empty:
            stops_bl["id_punto_opt"] = (
                stops_bl[cols["col_reserva"]].astype(str)
                if cols["col_reserva"]
                else "-"
            )
            stops_bl["vehiculo_base"] = stops_bl["ruta"]
            stops_bl["eta"] = (
                stops_bl["datetime_gestion"]
                .dt.strftime("%H:%M")
            )
            stops_bl["os"] = (
                pd.to_numeric(
                    stops_bl[cols["col_q_os"]],
                    errors="coerce"
                ).fillna(0)
                if cols["col_q_os"]
                else 1
            )

            if cols["col_km"]:
                q_os_mapa = (
                    pd.to_numeric(
                        stops_bl[cols["col_q_os"]],
                        errors="coerce"
                    ).fillna(1)
                    if cols["col_q_os"]
                    else pd.Series(
                        1,
                        index=stops_bl.index
                    )
                )

                ajuste_mapa = ajustar_volumen_m3(
                    stops_bl[cols["col_km"]],
                    q_os_mapa,
                    ajuste_vol_por_os=(
                        stops_bl["ruta"].map(
                            dicruta.set_index("ruta")[
                                "ajuste_vol"
                            ].to_dict()
                        )
                        if "ajuste_vol" in dicruta.columns
                        else None
                    ),
                    umbral_m3_por_os=3.0,
                    m3_estandar_por_os=0.33
                )

                stops_bl["volumen_original_m3"] = (
                    ajuste_mapa["volumen_original_m3"]
                )
                stops_bl["volumen"] = (
                    ajuste_mapa["volumen_m3"]
                )
                stops_bl["m3_por_os"] = (
                    ajuste_mapa["m3_por_os"]
                )
                stops_bl["volumen_ajustado"] = (
                    ajuste_mapa["volumen_ajustado"]
                )
            else:
                stops_bl["volumen_original_m3"] = 0
                stops_bl["volumen"] = 0
                stops_bl["m3_por_os"] = 0
                stops_bl["volumen_ajustado"] = False

        render_map(
            stops_bl,
            paths_bl,
            "Mapa Baseline por calles OSRM",
            key_prefix="mapa_baseline"
        )

    with sub2:
        stops_c, paths_c = preparar_mapa_opt(
            sim["routes_cap"]
        )
        render_map(
            stops_c,
            paths_c,
            "Mapa Optimización de capacidades por calles OSRM",
            key_prefix="mapa_opt_cap"
        )

    with sub3:
        st.markdown("#### Resumen por ruta / vuelta")

        resumen_bl_ruta = resumen_por_ruta_baseline(
            sim["baseline"],
            dicruta,
            cols
        )
        resumen_bl_vuelta = resumen_por_vuelta_baseline(
            sim["baseline"],
            dicruta,
            cols
        )

        t1, t2, t3 = st.tabs([
            "Baseline ruta",
            "Baseline vuelta",
            "Opt. capacidades"
        ])

        with t1:
            st.dataframe(
                resumen_bl_ruta,
                use_container_width=True
            )

        with t2:
            st.dataframe(
                resumen_bl_vuelta,
                use_container_width=True
            )

        with t3:
            st.dataframe(
                sim["resumen_cap_opt"],
                use_container_width=True
            )

    with sub4:
        st.markdown("#### ¿Qué rutas están mejor y peor utilizadas?")
        st.caption(
            "Ordenado por % de ocupación. Las rutas al final de cada lista son las que más "
            "capacidad ociosa dejan en el día — primeras candidatas a consolidar o eliminar."
        )

        rank_bl = resumen_por_ruta_baseline(sim["baseline"], dicruta, cols)[
            ["ruta", "volumen", "capacidad", "factor_ocupacion", "vueltas", "reservas"]
        ].dropna(subset=["factor_ocupacion"]).sort_values("factor_ocupacion", ascending=False)

        rank_cap = sim["resumen_cap_opt"].copy()
        if not rank_cap.empty and "factor_ocupacion" in rank_cap.columns:
            rank_cap = rank_cap[
                ["vehiculo_base", "bloque", "vuelta", "volumen", "capacidad", "factor_ocupacion", "paradas"]
            ].dropna(subset=["factor_ocupacion"]).sort_values("factor_ocupacion", ascending=False)

        rk1, rk2 = st.columns(2)
        with rk1:
            st.markdown("##### Baseline")
            if rank_bl.empty:
                st.info("Sin datos de ocupación por ruta para este día.")
            else:
                st.caption(f"🟢 Mejor ocupadas ({min(3, len(rank_bl))})")
                st.dataframe(rank_bl.head(3), use_container_width=True, hide_index=True)
                st.caption(f"🔴 Con más capacidad ociosa ({min(3, len(rank_bl))})")
                st.dataframe(rank_bl.tail(3).sort_values("factor_ocupacion"), use_container_width=True, hide_index=True)

                chart_bl = alt.Chart(rank_bl).mark_bar().encode(
                    x=alt.X("factor_ocupacion:Q", title="% Ocupación"),
                    y=alt.Y("ruta:N", sort="-x", title="Ruta"),
                    color=alt.condition("datum.factor_ocupacion < 30", alt.value("#F04438"), alt.value("#12B76A")),
                    tooltip=["ruta:N", alt.Tooltip("factor_ocupacion:Q", format=".1f"), alt.Tooltip("volumen:Q", format=".2f")]
                ).properties(height=max(200, 22 * len(rank_bl)))
                st.altair_chart(chart_bl, use_container_width=True)

        with rk2:
            st.markdown("##### Optimización de capacidades")
            if rank_cap.empty:
                st.info("Sin datos de ocupación por vehículo para este día.")
            else:
                st.caption(f"🟢 Mejor ocupadas ({min(3, len(rank_cap))})")
                st.dataframe(rank_cap.head(3), use_container_width=True, hide_index=True)
                st.caption(f"🔴 Con más capacidad ociosa ({min(3, len(rank_cap))})")
                st.dataframe(rank_cap.tail(3).sort_values("factor_ocupacion"), use_container_width=True, hide_index=True)

                chart_cap = alt.Chart(rank_cap).mark_bar().encode(
                    x=alt.X("factor_ocupacion:Q", title="% Ocupación"),
                    y=alt.Y("vehiculo_base:N", sort="-x", title="Vehículo / ruta"),
                    color=alt.condition("datum.factor_ocupacion < 30", alt.value("#F04438"), alt.value("#12B76A")),
                    tooltip=["vehiculo_base:N", "bloque:N", alt.Tooltip("factor_ocupacion:Q", format=".1f"), alt.Tooltip("volumen:Q", format=".2f")]
                ).properties(height=max(200, 22 * len(rank_cap)))
                st.altair_chart(chart_cap, use_container_width=True)

with tab_exportar:
    st.markdown('<div class="section-title">Exportar</div>', unsafe_allow_html=True)
    st.caption("Ahora puedes descargar uno o varios días, no solo el día que estás viendo en pantalla.")

    clases_exp = filtro_clases
    tipos_exp = filtro_tipos
    comuna_exp = filtro_comuna
    st.info(f"Exportar usará los filtros generales aplicados · Comuna: {comuna_exp}")

    peu_fechas_exp = peu[(peu["clase_ruta"].isin(clases_exp)) & (peu["tipo_ruta"].isin(tipos_exp))].copy()
    if comuna_exp and comuna_exp != "Todas":
        peu_fechas_exp = peu_fechas_exp[peu_fechas_exp["comuna"].eq(comuna_exp)]
    fechas_exp = sorted(peu_fechas_exp["fecha_operacion"].dropna().unique())
    fechas_sel_exp = st.multiselect("Días a descargar", fechas_exp, default=fechas_exp[:1], format_func=lambda x: str(x))

    if not fechas_sel_exp:
        st.warning("Selecciona al menos un día para descargar.")
    else:
        st.info(f"Se exportarán {len(fechas_sel_exp)} día(s).")

        if st.button("Preparar Excel de descarga", type="primary"):
            sims=[]
            prog = st.progress(0, text="Preparando descarga...")
            for i, f in enumerate(fechas_sel_exp):
                sim_tmp = ejecutar_escenarios_dia(
                    f, clases_exp, tipos_exp, comuna_exp,
                    optimizar_reservas_propias=optimizar_reservas_propias,
                    vehiculos_propios_incluidos=vehiculos_propios_incluidos,
            tipos_ruta_priorizados=tipos_ruta_priorizados
                )
                sims.append((f, sim_tmp))
                prog.progress((i+1)/len(fechas_sel_exp), text=f"Días preparados: {i+1}/{len(fechas_sel_exp)}")
            prog.empty()

            output_name = "baseline_optimizador_capacidades_export.xlsx"
            all_metrics=[]; all_baseline=[]; all_bl_ruta=[]; all_bl_vuelta=[]
            all_opt_cap_res=[]; all_opt_cap_det=[]; all_puntos=[]; all_no_asig=[]; all_diag_bl=[]; all_tipos_ruta=[]

            for f, sim_tmp in sims:
                all_metrics.append(sim_tmp["metrics"].assign(fecha_export=f))
                all_baseline.append(sim_tmp["baseline"].assign(fecha_export=f))
                all_bl_ruta.append(resumen_por_ruta_baseline(sim_tmp["baseline"], dicruta, cols).assign(fecha_export=f))
                all_bl_vuelta.append(resumen_por_vuelta_baseline(sim_tmp["baseline"], dicruta, cols).assign(fecha_export=f))
                all_opt_cap_res.append(sim_tmp["resumen_cap_opt"].assign(fecha_export=f))
                all_opt_cap_det.append(sim_tmp["detail_cap"].assign(fecha_export=f))
                all_puntos.append(sim_tmp["puntos"].assign(fecha_export=f))
                all_no_asig.append(sim_tmp.get("no_asignadas_cap", pd.DataFrame()).assign(fecha_export=f))
                all_diag_bl.append(sim_tmp.get("diagnostico_baseline", pd.DataFrame()).assign(fecha_export=f))
                all_tipos_ruta.append(sim_tmp.get("comparacion_tipos_ruta", pd.DataFrame()).assign(fecha_export=f))

            resumen_ejecutivo_export = construir_resumen_ejecutivo_export(sims)

            output_buffer = BytesIO()
            with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
                resumen_ejecutivo_export.to_excel(writer, sheet_name="resumen_ejecutivo", index=False)
                concatenar_no_vacios(all_metrics).to_excel(writer, sheet_name="macro_escenarios", index=False)
                concatenar_no_vacios(all_tipos_ruta).to_excel(writer, sheet_name="comparacion_tipos_ruta", index=False)
                concatenar_no_vacios(all_baseline).to_excel(writer, sheet_name="baseline_detalle", index=False)
                concatenar_no_vacios(all_bl_ruta).to_excel(writer, sheet_name="baseline_resumen_ruta", index=False)
                concatenar_no_vacios(all_bl_vuelta).to_excel(writer, sheet_name="baseline_resumen_vuelta", index=False)
                concatenar_no_vacios(all_diag_bl).to_excel(writer, sheet_name="baseline_diagnostico", index=False)
                concatenar_no_vacios(all_opt_cap_res).to_excel(writer, sheet_name="opt_cap_resumen", index=False)
                concatenar_no_vacios(all_opt_cap_det).to_excel(writer, sheet_name="opt_cap_detalle", index=False)
                concatenar_no_vacios(all_puntos).to_excel(writer, sheet_name="demanda_optimizable", index=False)
                concatenar_no_vacios(all_no_asig).to_excel(writer, sheet_name="reservas_no_asignadas", index=False)
                pd.DataFrame(DICRUTA_VALIDACIONES).to_excel(writer, sheet_name="validaciones_dicruta", index=False)
                dicruta.to_excel(writer, sheet_name="dicruta_configuracion", index=False)
                pd.DataFrame(GLOSARIO_EXPORT).to_excel(writer, sheet_name="glosario", index=False)

                formatear_libro_excel(writer.book)

            output_buffer.seek(0)
            st.download_button(
                "Descargar Excel completo",
                data=output_buffer.getvalue(),
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
